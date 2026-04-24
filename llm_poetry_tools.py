"""LLM-based editing, evaluation, and reporting helpers for the poetry project."""

import csv
import json
import os
import re
import statistics
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence


try:
    from google import genai
    from google.genai import types
except Exception:  # pragma: no cover - optional dependency in non-Colab runs.
    genai = None
    types = None

try:
    import torch
except Exception:  # pragma: no cover - optional dependency in non-Colab runs.
    torch = None

try:
    from transformers import AutoModelForCausalLM, AutoTokenizer, BitsAndBytesConfig
except Exception:  # pragma: no cover - optional dependency in non-Colab runs.
    AutoModelForCausalLM = None
    AutoTokenizer = None
    BitsAndBytesConfig = None

try:
    import httpx
except Exception:  # pragma: no cover - optional dependency in non-Colab runs.
    httpx = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - optional dependency in non-Colab runs.
    Workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None


VOWELS = "аеёиоуыэюя"
DEFAULT_GEMINI_MODEL = "gemini-2.5-flash-lite"
DEFAULT_OLLAMA_MODEL = "qwen2.5:7b-instruct"
DEFAULT_TRANSFORMERS_MODEL = "mistralai/Mistral-7B-Instruct-v0.3"
DEFAULT_GROQ_MODEL = "llama-3.3-70b-versatile"
GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"
GROQ_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/136.0.0.0 Safari/537.36 PoetryDesktop/1.0"
)
SCORE_FIELDS = [
    "semantic_coherence",
    "grammar",
    "theme_consistency",
    "poetic_quality",
    "overall",
]
REPORT_NUMERIC_FIELDS = [
    "requested_lines",
    "actual_lines",
    "word_count_total",
    "char_count_total",
    "unique_line_count",
    "rhyme_success",
    "rhyme_quality",
    "unique_rate",
    "unique_words",
    "avg_words_per_line",
    "avg_vowels_per_line",
    *SCORE_FIELDS,
]


@dataclass
class LLMConfig:
    """Configuration for the optional LLM editor/evaluator."""

    provider: str = "gemini"
    model_name: str = DEFAULT_GEMINI_MODEL
    editor_temperature: float = 0.7
    evaluator_temperature: float = 0.0
    max_retries: int = 2
    retry_delay_seconds: float = 2.0
    request_timeout_seconds: float = 90.0
    ollama_base_url: str = "http://127.0.0.1:11434"
    transformers_max_new_tokens: int = 256


def _get_colab_secret(name: str) -> Optional[str]:
    """Read a secret from Google Colab if the notebook is running there."""

    try:
        from google.colab import userdata  # type: ignore

        value = userdata.get(name)
        return value or None
    except Exception:
        return None


def get_gemini_api_key() -> Optional[str]:
    """Find a Gemini API key in environment variables or Colab Secrets."""

    return (
        os.getenv("GEMINI_API_KEY")
        or os.getenv("GOOGLE_API_KEY")
        or _get_colab_secret("GEMINI_API_KEY")
        or _get_colab_secret("GOOGLE_API_KEY")
    )


def get_llm_provider() -> Optional[str]:
    """Find the selected LLM provider in environment variables or Colab Secrets."""

    return os.getenv("LLM_PROVIDER") or _get_colab_secret("LLM_PROVIDER")


def get_groq_api_key() -> Optional[str]:
    """Find a Groq API key in environment variables or Colab Secrets."""

    return os.getenv("GROQ_API_KEY") or _get_colab_secret("GROQ_API_KEY")


def get_ollama_base_url() -> Optional[str]:
    """Find the Ollama base URL in environment variables or Colab Secrets."""

    return (
        os.getenv("OLLAMA_BASE_URL")
        or _get_colab_secret("OLLAMA_BASE_URL")
        or os.getenv("OLLAMA_HOST")
        or _get_colab_secret("OLLAMA_HOST")
    )


def tokenize_russian(text: str) -> List[str]:
    """Extract Russian words from text."""

    return re.findall(r"[а-яё]+", text.lower())


def normalize_poem(poem: Sequence[str] | str | None) -> List[str]:
    """Convert a poem represented as text or list of lines to clean lines."""

    if poem is None:
        return []

    if isinstance(poem, str):
        raw_lines = poem.splitlines()
    else:
        raw_lines = list(poem)

    lines = []
    for line in raw_lines:
        clean = str(line).strip()
        if not clean:
            continue
        clean = re.sub(r"^\s*[\-\*\d]+[\.\)\:\-]?\s*", "", clean).strip()
        if clean:
            lines.append(clean)
    return lines


def poem_to_text(poem: Sequence[str] | str | None) -> str:
    """Return a poem as newline-separated text."""

    return "\n".join(normalize_poem(poem))


def print_poem(title: str, poem: Sequence[str] | str | None) -> None:
    """Print a poem with a compact title."""

    print(f"\n{title}\n")
    lines = normalize_poem(poem)
    if not lines:
        print("Стихотворение отсутствует")
        return
    for line in lines:
        print(line)


def simplify_rhyme_scheme(scheme: str, length: int) -> List[str]:
    """Repeat or trim a rhyme scheme to match the poem length."""

    clean = "".join(ch.upper() for ch in str(scheme) if ch.isalpha()) or "AABB"
    if len(clean) < length:
        clean = (clean * ((length + len(clean) - 1) // len(clean)))[:length]
    return list(clean[:length])


def extract_last_word(line: str) -> Optional[str]:
    """Return the last Russian word from a line."""

    words = tokenize_russian(line)
    return words[-1] if words else None


def calculate_rhyme_quality(line1: str, line2: str) -> float:
    """Score a rhyme by matching suffixes of the last words."""

    last1 = extract_last_word(line1)
    last2 = extract_last_word(line2)
    if not last1 or not last2:
        return 0.0
    if len(last1) < 3 or len(last2) < 3:
        return 0.5

    for suffix_len in range(3, 0, -1):
        if last1[-suffix_len:] == last2[-suffix_len:]:
            return {3: 1.0, 2: 0.8, 1: 0.5}[suffix_len]
    return 0.2


def calculate_formal_metrics(
    poem: Sequence[str] | str | None,
    rhyme_scheme: str = "AABB",
) -> Dict[str, Any]:
    """Calculate simple comparable metrics for raw or LLM-edited poems."""

    lines = normalize_poem(poem)
    if not lines:
        return {
            "lines": 0,
            "rhyme_success": 0.0,
            "rhyme_quality": 0.0,
            "unique_rate": 0.0,
            "unique_words": 0,
            "avg_words_per_line": 0.0,
            "avg_vowels_per_line": 0.0,
        }

    scheme = simplify_rhyme_scheme(rhyme_scheme, len(lines))
    last_line_by_letter: Dict[str, str] = {}
    rhyme_scores = []

    for line, letter in zip(lines, scheme):
        previous_line = last_line_by_letter.get(letter)
        if previous_line:
            rhyme_scores.append(calculate_rhyme_quality(previous_line, line))
        last_line_by_letter[letter] = line

    tokenized = [tokenize_russian(line) for line in lines]
    all_words = [word for words in tokenized for word in words]
    vowels_per_line = [
        sum(1 for char in line.lower() if char in VOWELS) for line in lines
    ]
    words_per_line = [len(words) for words in tokenized]

    unique_lines = len({line.lower() for line in lines})
    successful_rhymes = sum(1 for score in rhyme_scores if score >= 0.5)

    return {
        "lines": len(lines),
        "rhyme_success": round(successful_rhymes / max(1, len(rhyme_scores)) * 100, 1),
        "rhyme_quality": round(
            sum(rhyme_scores) / max(1, len(rhyme_scores)) * 100, 1
        ),
        "unique_rate": round(unique_lines / len(lines) * 100, 1),
        "unique_words": len(set(all_words)),
        "avg_words_per_line": round(sum(words_per_line) / len(words_per_line), 1),
        "avg_vowels_per_line": round(sum(vowels_per_line) / len(vowels_per_line), 1),
    }


def build_formal_metrics_table(
    poem_versions: Dict[str, Sequence[str] | str | None],
    rhyme_scheme: str = "AABB",
) -> List[Dict[str, Any]]:
    """Build rows for a formal metrics comparison table."""

    rows = []
    for name, poem in poem_versions.items():
        row = {"name": name}
        row.update(calculate_formal_metrics(poem, rhyme_scheme=rhyme_scheme))
        rows.append(row)
    return rows


def poem_lines_to_columns(
    poem: Sequence[str] | str | None,
    max_lines: int = 16,
) -> Dict[str, str]:
    """Expand poem lines into fixed CSV-friendly columns."""

    lines = normalize_poem(poem)
    columns = {}
    for index in range(max_lines):
        key = f"line_{index + 1}"
        columns[key] = lines[index] if index < len(lines) else ""
    return columns


def detect_source_model(version_name: str) -> str:
    """Infer the base model from a version label."""

    label = str(version_name)
    if "Markov" in label:
        return "Markov"
    if "LSTM" in label:
        return "LSTM"
    return ""


def build_poem_report_row(
    version: str,
    poem: Sequence[str] | str | None,
    report_scope: str = "batch",
    run: Optional[int] = None,
    source_model: Optional[str] = None,
    stage: Optional[str] = None,
    start_line: str = "",
    rhyme_scheme: str = "AABB",
    requested_lines: Optional[int] = None,
    llm_model_name: str = "",
    metrics: Optional[Dict[str, Any]] = None,
    evaluation: Optional[Dict[str, Any]] = None,
    max_lines: int = 16,
    extra: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Build one detailed CSV row for a poem version."""

    lines = normalize_poem(poem)
    text = poem_to_text(lines)
    metrics = metrics or calculate_formal_metrics(lines, rhyme_scheme=rhyme_scheme)
    evaluation = evaluation or {}
    source_model = source_model or detect_source_model(version)
    stage = stage or ("llm_edited" if "+ LLM" in str(version) else "raw")
    words = tokenize_russian(text)

    row: Dict[str, Any] = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "report_scope": report_scope,
        "run": run if run is not None else "",
        "source_model": source_model,
        "version": version,
        "stage": stage,
        "llm_used": "yes" if stage == "llm_edited" else "no",
        "llm_model_name": llm_model_name if llm_model_name else "",
        "start_line": start_line.strip() if start_line else "",
        "rhyme_scheme": rhyme_scheme,
        "requested_lines": requested_lines if requested_lines is not None else len(lines),
        "actual_lines": len(lines),
        "unique_line_count": len({line.lower() for line in lines}),
        "word_count_total": len(words),
        "char_count_total": len(text),
        "poem_text": text,
    }
    row.update(poem_lines_to_columns(lines, max_lines=max_lines))
    row.update(metrics)

    for field in SCORE_FIELDS:
        row[field] = evaluation.get(field, "")
    row["llm_comment"] = evaluation.get("comment", "")
    row["llm_raw_response"] = evaluation.get("raw_response", "")

    if extra:
        row.update(extra)

    return row


def summarize_report_rows(
    rows: Sequence[Dict[str, Any]],
    group_fields: Sequence[str] = ("report_scope", "version", "stage"),
) -> List[Dict[str, Any]]:
    """Aggregate detailed report rows into a compact summary table."""

    grouped: Dict[tuple, List[Dict[str, Any]]] = {}
    for row in rows:
        key = tuple(row.get(field, "") for field in group_fields)
        grouped.setdefault(key, []).append(row)

    summary_rows = []
    for key, group in grouped.items():
        summary = {field: value for field, value in zip(group_fields, key)}
        summary["row_count"] = len(group)
        summary["source_model"] = group[0].get("source_model", "")
        summary["llm_used"] = group[0].get("llm_used", "")
        summary["llm_model_name"] = group[0].get("llm_model_name", "")

        for field in REPORT_NUMERIC_FIELDS:
            values = []
            for row in group:
                value = row.get(field, "")
                if value in ("", None):
                    continue
                try:
                    values.append(float(value))
                except Exception:
                    continue
            summary[f"mean_{field}"] = (
                round(statistics.mean(values), 3) if values else ""
            )

        comments = [str(row.get("llm_comment", "")).strip() for row in group]
        comments = [comment for comment in comments if comment]
        summary["sample_comment"] = comments[0] if comments else ""
        summary_rows.append(summary)

    return summary_rows


def write_csv_report(
    rows: Sequence[Dict[str, Any]],
    output_path: str | Path,
    preferred_fields: Optional[Sequence[str]] = None,
) -> str:
    """Write rows to CSV and return the resulting path."""

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = resolve_report_fieldnames(rows, preferred_fields=preferred_fields)

    with path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})

    return str(path)


def resolve_report_fieldnames(
    rows: Sequence[Dict[str, Any]],
    preferred_fields: Optional[Sequence[str]] = None,
) -> List[str]:
    """Build a stable ordered field list for report export."""

    preferred_fields = list(preferred_fields or [])
    keys = []
    for row in rows:
        for key in row.keys():
            if key not in keys:
                keys.append(key)

    extra_fields = [key for key in keys if key not in preferred_fields]
    return preferred_fields + extra_fields


def _excel_cell_value(value: Any) -> Any:
    """Convert Python values to something spreadsheet-friendly."""

    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False)


def _column_width_for_field(fieldname: str, values: Sequence[Any]) -> float:
    """Pick a practical Excel column width."""

    lowered = fieldname.lower()
    if lowered == "poem_text":
        return 42
    if lowered in {"llm_comment", "llm_raw_response"}:
        return 36
    if lowered.startswith("line_"):
        return 28
    if lowered.endswith("_path"):
        return 32

    sample_lengths = [len(str(value)) for value in values if value not in ("", None)]
    header_len = len(fieldname)
    if sample_lengths:
        sample_lengths.sort()
        percentile_index = int((len(sample_lengths) - 1) * 0.75)
        typical_len = sample_lengths[percentile_index]
        return max(12, min(24, max(header_len, typical_len + 2)))
    return max(12, min(20, header_len + 2))


def write_excel_report(
    rows: Sequence[Dict[str, Any]],
    output_path: str | Path,
    preferred_fields: Optional[Sequence[str]] = None,
    sheet_name: str = "Report",
) -> Optional[str]:
    """Write rows to a single-sheet Excel workbook and return the path."""

    if Workbook is None or get_column_letter is None:
        return None

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = resolve_report_fieldnames(rows, preferred_fields=preferred_fields)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (sheet_name or "Report")[:31]

    sheet.append(fieldnames)
    for row in rows:
        sheet.append([_excel_cell_value(row.get(field, "")) for field in fieldnames])

    if Font is not None and PatternFill is not None and Alignment is not None:
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        body_alignment = Alignment(vertical="top", wrap_text=True)

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = body_alignment

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for index, fieldname in enumerate(fieldnames, start=1):
        values = [row.get(fieldname, "") for row in rows]
        sheet.column_dimensions[get_column_letter(index)].width = _column_width_for_field(
            fieldname, values
        )

    workbook.save(path)
    return str(path)


def export_report_bundle(
    rows: Sequence[Dict[str, Any]],
    output_dir: str | Path,
    prefix: str = "poetry_report",
    summary_group_fields: Sequence[str] = ("report_scope", "version", "stage"),
) -> Dict[str, Any]:
    """Export detailed and summary reports in user-friendly and fallback formats."""

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    preferred_detail_fields = [
        "created_at",
        "report_scope",
        "run",
        "source_model",
        "version",
        "stage",
        "llm_used",
        "llm_model_name",
        "start_line",
        "rhyme_scheme",
        "requested_lines",
        "actual_lines",
        "unique_line_count",
        "word_count_total",
        "char_count_total",
        "rhyme_success",
        "rhyme_quality",
        "unique_rate",
        "unique_words",
        "avg_words_per_line",
        "avg_vowels_per_line",
        *SCORE_FIELDS,
        "llm_comment",
        "llm_raw_response",
        "poem_text",
    ] + [f"line_{index}" for index in range(1, 17)]

    detailed_rows = list(rows)
    summary_rows = summarize_report_rows(
        detailed_rows, group_fields=summary_group_fields
    )
    detailed_csv_path = output_dir / f"{prefix}_detailed_{timestamp}.csv"
    summary_csv_path = output_dir / f"{prefix}_summary_{timestamp}.csv"
    detailed_xlsx_path = output_dir / f"{prefix}_detailed_{timestamp}.xlsx"
    summary_xlsx_path = output_dir / f"{prefix}_summary_{timestamp}.xlsx"

    write_csv_report(
        detailed_rows,
        detailed_csv_path,
        preferred_fields=preferred_detail_fields,
    )
    write_csv_report(summary_rows, summary_csv_path)

    detailed_path = write_excel_report(
        detailed_rows,
        detailed_xlsx_path,
        preferred_fields=preferred_detail_fields,
        sheet_name="Detailed",
    )
    summary_path = write_excel_report(
        summary_rows,
        summary_xlsx_path,
        sheet_name="Summary",
    )

    return {
        "detailed_path": detailed_path or str(detailed_csv_path),
        "summary_path": summary_path or str(summary_csv_path),
        "detailed_csv_path": str(detailed_csv_path),
        "summary_csv_path": str(summary_csv_path),
        "detailed_xlsx_path": str(detailed_xlsx_path) if detailed_path else "",
        "summary_xlsx_path": str(summary_xlsx_path) if summary_path else "",
        "detailed_rows": detailed_rows,
        "summary_rows": summary_rows,
    }


class LLMPoetryAssistant:
    """Optional LLM-based editor and evaluator."""

    def __init__(
        self,
        api_key: Optional[str] = None,
        config: Optional[LLMConfig] = None,
        model_name: Optional[str] = None,
        provider: Optional[str] = None,
        base_url: Optional[str] = None,
    ):
        self.config = config or LLMConfig()
        provider_value = provider or get_llm_provider() or self.config.provider
        self.provider = str(provider_value).strip().lower() or "gemini"
        if model_name:
            self.config.model_name = model_name

        self.client = None
        self.tokenizer = None
        self.base_url: Optional[str] = None
        self.device_label: str = "cpu"
        self.last_error: Optional[str] = None

        if self.provider in {"hf", "huggingface"}:
            self.provider = "transformers"

        if self.provider in {"disabled", "none", "off"}:
            self.provider = "disabled"
            self.last_error = "LLM отключен текущими настройками."
            return

        if self.provider not in {"gemini", "groq", "ollama", "transformers"}:
            self.last_error = (
                f"Unsupported LLM provider: {self.provider}. "
                "Use 'groq', 'gemini', 'ollama', 'transformers', or 'disabled'."
            )
            return

        # Resolve API key after provider is known so we look in the right place.
        if self.provider == "groq":
            self.api_key = api_key or get_groq_api_key()
        else:
            self.api_key = api_key or get_gemini_api_key()

        if self.provider == "groq":
            self._init_groq()
            return

        if self.provider == "ollama":
            self._init_ollama(base_url=base_url)
            return

        if self.provider == "transformers":
            self._init_transformers()
            return

        if genai is None:
            self.last_error = "Не установлена библиотека google-genai."
            return

        if not self.api_key:
            self.last_error = (
                "Не найден GEMINI_API_KEY. Добавьте ключ в переменную окружения "
                "или, если вы всё же запускаете ноутбук в Colab, в Colab Secrets."
            )
            return

        try:
            self.client = genai.Client(api_key=self.api_key)
        except Exception as exc:
            self.last_error = f"Не удалось создать Gemini client: {exc}"

    def _init_ollama(self, base_url: Optional[str] = None) -> None:
        if not self.config.model_name or self.config.model_name == DEFAULT_GEMINI_MODEL:
            self.config.model_name = DEFAULT_OLLAMA_MODEL

        ollama_base_url = base_url or get_ollama_base_url() or self.config.ollama_base_url
        self.base_url = str(ollama_base_url).rstrip("/") if ollama_base_url else None

        if not self.base_url:
            self.last_error = (
                "OLLAMA_BASE_URL is not set. Point it to your local Ollama server "
                "or to a tunnel URL if you access Ollama remotely."
            )
            return

        self.client = "ollama"

    def _init_groq(self) -> None:
        if not self.config.model_name or self.config.model_name in {
            DEFAULT_GEMINI_MODEL,
            DEFAULT_OLLAMA_MODEL,
        }:
            self.config.model_name = DEFAULT_GROQ_MODEL

        if not self.api_key:
            self.last_error = (
                "Не найден GROQ_API_KEY. "
                "Получите бесплатный ключ на console.groq.com и добавьте "
                "переменную окружения GROQ_API_KEY."
            )
            return

        self.client = "groq"

    def _generate_groq(
        self,
        prompt: str,
        temperature: float,
        response_mime_type: Optional[str] = None,
    ) -> str:
        if not self.api_key:
            raise RuntimeError(self.status_message())

        payload: Dict[str, Any] = {
            "model": self.config.model_name,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": max(0.0, float(temperature)),
            "max_tokens": self.config.transformers_max_new_tokens,
        }
        if response_mime_type == "application/json":
            payload["response_format"] = {"type": "json_object"}

        last_exc: Optional[Exception] = None
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
            "Authorization": f"Bearer {self.api_key}",
            "User-Agent": GROQ_USER_AGENT,
        }

        for attempt in range(self.config.max_retries + 1):
            try:
                if httpx is not None:
                    with httpx.Client(
                        timeout=self.config.request_timeout_seconds,
                        follow_redirects=True,
                    ) as client:
                        response = client.post(
                            GROQ_API_URL,
                            json=payload,
                            headers=headers,
                        )
                    response.raise_for_status()
                    data = response.json()
                else:
                    request = urllib.request.Request(
                        GROQ_API_URL,
                        data=json.dumps(payload).encode("utf-8"),
                        headers=headers,
                        method="POST",
                    )
                    with urllib.request.urlopen(
                        request,
                        timeout=self.config.request_timeout_seconds,
                    ) as response:
                        data = json.loads(response.read().decode("utf-8"))

                return data["choices"][0]["message"]["content"].strip()
            except Exception as exc:
                if httpx is not None and isinstance(exc, httpx.HTTPStatusError):
                    error_body = exc.response.text
                    if exc.response.status_code == 403 and "1010" in error_body:
                        last_exc = RuntimeError(
                            "Groq HTTP 403 / Cloudflare 1010. "
                            "Запрос заблокирован Browser Integrity Check или WAF. "
                            "Ключ может быть корректным, но текущий IP/сигнатура клиента отклонены."
                        )
                    else:
                        last_exc = RuntimeError(
                            f"Groq HTTP {exc.response.status_code}: {error_body}"
                        )
                elif httpx is not None and isinstance(exc, httpx.HTTPError):
                    last_exc = RuntimeError(f"Groq HTTP client error: {exc}")
                elif isinstance(exc, urllib.error.HTTPError):
                    error_body = exc.read().decode("utf-8", errors="replace")
                    if exc.code == 403 and "1010" in error_body:
                        last_exc = RuntimeError(
                            "Groq HTTP 403 / Cloudflare 1010. "
                            "Запрос заблокирован Browser Integrity Check или WAF. "
                            "Ключ может быть корректным, но текущий IP/сигнатура клиента отклонены."
                        )
                    else:
                        last_exc = RuntimeError(
                            f"Groq HTTP {exc.code}: {error_body or exc.reason}"
                        )
                else:
                    last_exc = exc

            if attempt < self.config.max_retries:
                time.sleep(self.config.retry_delay_seconds * (attempt + 1))

        raise RuntimeError(f"Groq API error: {last_exc}")

    def _init_transformers(self) -> None:
        if AutoModelForCausalLM is None or AutoTokenizer is None or torch is None:
            self.last_error = (
                "Не установлены transformers/torch. "
                "Установите transformers, accelerate и torch в локальное окружение."
            )
            return

        if not self.config.model_name or self.config.model_name in {
            DEFAULT_GEMINI_MODEL,
            DEFAULT_OLLAMA_MODEL,
        }:
            self.config.model_name = DEFAULT_TRANSFORMERS_MODEL

        if not torch.cuda.is_available():
            self.last_error = (
                "Для локального transformers-LLM нужна CUDA GPU. "
                "Используйте локальное CUDA-окружение или переключите "
                "LLM_PROVIDER на 'disabled' или 'ollama'."
            )
            return

        model_kwargs: Dict[str, Any] = {
            "low_cpu_mem_usage": True,
            "torch_dtype": torch.float16,
            "device_map": "auto",
        }
        self.device_label = "cuda"
        if BitsAndBytesConfig is not None:
            model_kwargs["quantization_config"] = BitsAndBytesConfig(
                load_in_4bit=True,
                bnb_4bit_quant_type="nf4",
                bnb_4bit_use_double_quant=True,
                bnb_4bit_compute_dtype=torch.float16,
            )

        try:
            self.tokenizer = AutoTokenizer.from_pretrained(self.config.model_name)
            self.client = AutoModelForCausalLM.from_pretrained(
                self.config.model_name,
                **model_kwargs,
            )
            self.client.eval()
            if self.tokenizer.pad_token_id is None:
                self.tokenizer.pad_token = self.tokenizer.eos_token
        except Exception as exc:
            self.client = None
            self.tokenizer = None
            self.last_error = f"Не удалось загрузить локальную модель transformers: {exc}"

    def is_available(self) -> bool:
        """Return True if the selected LLM backend is ready."""

        if self.provider == "disabled":
            return False
        if self.provider == "ollama":
            return bool(self.base_url and self.client)
        if self.provider == "groq":
            return self.client == "groq" and bool(self.api_key)
        return self.client is not None

    def status_message(self) -> str:
        """Return a user-facing status line."""

        if self.is_available() and self.provider == "groq":
            return f"LLM-модуль готов: Groq / {self.config.model_name}"

        if self.is_available() and self.provider == "ollama":
            return (
                f"LLM-модуль готов: Ollama / {self.config.model_name} / {self.base_url}"
            )

        if self.is_available() and self.provider == "transformers":
            return (
                "LLM-модуль готов: "
                f"Transformers / {self.config.model_name} / {self.device_label}"
            )

        if self.is_available():
            return f"LLM-модуль готов: {self.config.model_name}"
        return self.last_error or "LLM-модуль недоступен."

    def test_connection(self) -> Dict[str, Any]:
        """Run a tiny request to verify that the current backend actually responds."""

        if not self.is_available():
            return {
                "ok": False,
                "provider": self.provider,
                "model_name": self.config.model_name,
                "message": self.status_message(),
            }

        prompt = "Reply with exactly OK"
        try:
            response_text = self._generate(
                prompt,
                temperature=0.0,
            ).strip()
            return {
                "ok": True,
                "provider": self.provider,
                "model_name": self.config.model_name,
                "message": self.status_message(),
                "response_text": response_text,
            }
        except Exception as exc:
            return {
                "ok": False,
                "provider": self.provider,
                "model_name": self.config.model_name,
                "message": f"{self.status_message()} | probe error: {exc}",
            }

    def _generate(
        self,
        prompt: str,
        temperature: float,
        response_mime_type: Optional[str] = None,
    ) -> str:
        if self.provider == "groq":
            return self._generate_groq(
                prompt,
                temperature=temperature,
                response_mime_type=response_mime_type,
            )

        if self.provider == "transformers":
            return self._generate_transformers(
                prompt,
                temperature=temperature,
                response_mime_type=response_mime_type,
            )

        if self.provider == "ollama":
            return self._generate_ollama(
                prompt,
                temperature=temperature,
                response_mime_type=response_mime_type,
            )

        if not self.client:
            raise RuntimeError(self.status_message())

        config_kwargs: Dict[str, Any] = {"temperature": temperature}
        if response_mime_type:
            config_kwargs["response_mime_type"] = response_mime_type

        generation_config = (
            types.GenerateContentConfig(**config_kwargs) if types else None
        )

        last_exc: Optional[Exception] = None
        for attempt in range(self.config.max_retries + 1):
            try:
                response = self.client.models.generate_content(
                    model=self.config.model_name,
                    contents=prompt,
                    config=generation_config,
                )
                return (response.text or "").strip()
            except Exception as exc:
                last_exc = exc
                if attempt < self.config.max_retries:
                    time.sleep(self.config.retry_delay_seconds * (attempt + 1))

        raise RuntimeError(f"Ошибка Gemini API: {last_exc}")

    def _generate_transformers(
        self,
        prompt: str,
        temperature: float,
        response_mime_type: Optional[str] = None,
    ) -> str:
        if not self.client or not self.tokenizer or torch is None:
            raise RuntimeError(self.status_message())

        messages = [{"role": "user", "content": prompt}]
        if hasattr(self.tokenizer, "apply_chat_template"):
            rendered_prompt = self.tokenizer.apply_chat_template(
                messages,
                tokenize=False,
                add_generation_prompt=True,
            )
        else:
            rendered_prompt = prompt

        encoded = self.tokenizer(
            rendered_prompt,
            return_tensors="pt",
            padding=False,
        )
        target_device = getattr(self.client, "device", None) or self.device_label
        encoded = {key: value.to(target_device) for key, value in encoded.items()}

        generate_kwargs: Dict[str, Any] = {
            "max_new_tokens": self.config.transformers_max_new_tokens,
            "pad_token_id": self.tokenizer.pad_token_id,
            "eos_token_id": self.tokenizer.eos_token_id,
        }
        if temperature > 0:
            generate_kwargs["do_sample"] = True
            generate_kwargs["temperature"] = max(0.05, float(temperature))
        else:
            generate_kwargs["do_sample"] = False

        with torch.no_grad():
            output = self.client.generate(
                **encoded,
                **generate_kwargs,
            )

        prompt_token_count = encoded["input_ids"].shape[1]
        generated_tokens = output[0][prompt_token_count:]
        generated_text = self.tokenizer.decode(
            generated_tokens,
            skip_special_tokens=True,
        ).strip()

        if response_mime_type == "application/json":
            match = re.search(r"\{.*\}", generated_text, flags=re.S)
            if match:
                generated_text = match.group(0)

        if torch.cuda.is_available():
            torch.cuda.empty_cache()

        return generated_text

    def _generate_ollama(
        self,
        prompt: str,
        temperature: float,
        response_mime_type: Optional[str] = None,
    ) -> str:
        if not self.base_url:
            raise RuntimeError(self.status_message())

        endpoint = f"{self.base_url}/api/generate"
        payload: Dict[str, Any] = {
            "model": self.config.model_name,
            "prompt": prompt,
            "stream": False,
            "options": {"temperature": temperature},
        }
        if response_mime_type == "application/json":
            payload["format"] = "json"

        last_exc: Optional[Exception] = None
        for attempt in range(self.config.max_retries + 1):
            try:
                request = urllib.request.Request(
                    endpoint,
                    data=json.dumps(payload).encode("utf-8"),
                    headers={"Content-Type": "application/json"},
                    method="POST",
                )
                with urllib.request.urlopen(
                    request,
                    timeout=self.config.request_timeout_seconds,
                ) as response:
                    response_data = json.loads(response.read().decode("utf-8"))
                return str(response_data.get("response", "")).strip()
            except urllib.error.HTTPError as exc:
                error_body = exc.read().decode("utf-8", errors="replace")
                last_exc = RuntimeError(
                    f"Ollama HTTP {exc.code}: {error_body or exc.reason}"
                )
            except Exception as exc:
                last_exc = exc

            if attempt < self.config.max_retries:
                time.sleep(self.config.retry_delay_seconds * (attempt + 1))

        raise RuntimeError(f"Ollama API error: {last_exc}")

    def edit_poem(
        self,
        draft_poem: Sequence[str] | str | None,
        start_line: str,
        rhyme_scheme: str = "AABB",
        preserve_first_line: bool = True,
        target_lines: Optional[int] = None,
    ) -> List[str]:
        """Rewrite a generated draft into a coherent poem using it as inspiration."""

        draft_lines = normalize_poem(draft_poem)
        if not draft_lines:
            return []

        requested_lines = max(0, int(target_lines or 0))
        expected_lines = max(len(draft_lines), requested_lines)
        draft_text = "\n".join(draft_lines)

        rhyme_scheme_explanation = {
            "AABB": "попарная: строки 1-2 рифмуются между собой, 3-4 рифмуются между собой, и т.д.",
            "ABAB": "перекрёстная: строки 1 и 3 рифмуются, строки 2 и 4 рифмуются",
            "ABBA": "опоясывающая: строки 1 и 4 рифмуются, строки 2 и 3 рифмуются",
            "AAAA": "сплошная: все строки рифмуются между собой",
        }.get(rhyme_scheme, rhyme_scheme)

        first_line_constraint = (
            f'Первая строка стихотворения должна быть точно: «{start_line}»'
            if preserve_first_line and start_line
            else ""
        )

        prompt = f"""
Ты русскоязычный поэт. Тебе дан черновик, созданный алгоритмом — он может быть бессвязным или грамматически сломанным. Это нормально.

Напиши осмысленное, художественное стихотворение на русском языке, вдохновлённое этим черновиком.

Жёсткие требования (нарушать нельзя):
1. Ровно {expected_lines} строк.
2. Схема рифмовки {rhyme_scheme}: {rhyme_scheme_explanation}. Проверь рифмы перед ответом.
{f"3. {first_line_constraint}" if first_line_constraint else ""}

Свобода действий:
- Можешь свободно переписать любые строки черновика или все строки целиком.
- Можешь менять слова, образы, порядок мыслей — главное, чтобы получился живой, связный текст.
- Черновик — только источник темы и настроения, не шаблон.

Качество результата:
- Текст должен быть грамматически правильным русским языком.
- Строки должны читаться как стихи, а не как набор слов.
- Рифмующиеся строки должны оканчиваться на одинаковые или близкие звуки (не просто похожие буквы).

Не добавляй заголовок, нумерацию строк, пояснения или комментарии.
Верни только готовое стихотворение.

Черновик (для вдохновения):
{draft_text}
""".strip()

        edited_text = self._generate(
            prompt,
            temperature=self.config.editor_temperature,
            response_mime_type="text/plain",
        )
        edited_lines = normalize_poem(edited_text)

        if not edited_lines:
            return draft_lines

        if len(edited_lines) > expected_lines:
            edited_lines = edited_lines[:expected_lines]
        elif len(edited_lines) < expected_lines:
            edited_lines.extend(
                self._complete_poem_lines(
                    current_lines=edited_lines,
                    draft_lines=draft_lines,
                    start_line=start_line,
                    rhyme_scheme=rhyme_scheme,
                    expected_lines=expected_lines,
                    preserve_first_line=preserve_first_line,
                )
            )

        if len(edited_lines) < expected_lines:
            edited_lines.extend(draft_lines[len(edited_lines) : expected_lines])

        if preserve_first_line and edited_lines and start_line:
            edited_lines[0] = start_line.strip()

        if len(edited_lines) > expected_lines:
            edited_lines = edited_lines[:expected_lines]

        return edited_lines

    def _complete_poem_lines(
        self,
        current_lines: Sequence[str],
        draft_lines: Sequence[str],
        start_line: str,
        rhyme_scheme: str,
        expected_lines: int,
        preserve_first_line: bool = True,
    ) -> List[str]:
        if len(current_lines) >= expected_lines:
            return []

        missing_lines = expected_lines - len(current_lines)
        current_text = "\n".join(current_lines)
        draft_text = "\n".join(draft_lines)
        first_line_rule = (
            f'- Первую строку сохрани без изменений: "{start_line}".'
            if preserve_first_line and start_line
            else "- Первую строку можно слегка отредактировать."
        )

        prompt = f"""
Ты завершаешь русскоязычное стихотворение.

Уже есть начало стихотворения, но оно неполное.
Добавь ровно {missing_lines} новых строки, чтобы в сумме получилось {expected_lines} строк.

Требования:
- Сохрани схему рифмовки: {rhyme_scheme}.
{first_line_rule}
- Сохрани тему и основные образы.
- Не повторяй уже имеющиеся строки.
- Верни только новые добавленные строки без нумерации и пояснений.

Уже готово:
{current_text}

Исходный черновик:
{draft_text}
""".strip()

        continuation_text = self._generate(
            prompt,
            temperature=self.config.editor_temperature,
            response_mime_type="text/plain",
        )
        continuation_lines = normalize_poem(continuation_text)
        return continuation_lines[:missing_lines]

    def evaluate_poem(
        self,
        poem: Sequence[str] | str | None,
        start_line: str,
    ) -> Dict[str, Any]:
        """Evaluate semantic quality with a fixed blind rubric."""

        poem_lines = normalize_poem(poem)
        if not poem_lines:
            return self._empty_evaluation("Стихотворение отсутствует.")

        poem_text = "\n".join(poem_lines)
        prompt = f"""
Ты независимый эксперт по оценке русскоязычных стихотворений.
Ты не знаешь, каким алгоритмом создан текст, и не должен это угадывать.

Оцени стихотворение по шкале от 1 до 5:
1 - очень плохо, 2 - плохо, 3 - средне, 4 - хорошо, 5 - отлично.

Используй всю шкалу 1-5, а не только крайние значения.
Ставь 1 только если текст почти полностью распадается по соответствующему критерию.
Не ставь всем критериям одинаковую оценку без необходимости: оцени их независимо.

Критерии:
- semantic_coherence: смысловая связность текста;
- grammar: грамматическая правильность;
- theme_consistency: сохранение темы первой строки;
- poetic_quality: художественность и выразительность;
- overall: общее качество.

Не оценивай строго классический стихотворный размер.
Не переписывай стихотворение.
Верни только JSON без markdown-разметки.

Первая строка:
"{start_line}"

Стихотворение:
{poem_text}

Формат ответа (подставь целые числа от 1 до 5 вместо угловых скобок):
{{
  "semantic_coherence": <число от 1 до 5>,
  "grammar": <число от 1 до 5>,
  "theme_consistency": <число от 1 до 5>,
  "poetic_quality": <число от 1 до 5>,
  "overall": <число от 1 до 5>,
  "comment": "краткий комментарий на русском"
}}
""".strip()

        raw_response = self._generate(
            prompt,
            temperature=self.config.evaluator_temperature,
            response_mime_type="application/json",
        )
        parsed = self._parse_evaluation(raw_response)
        parsed["raw_response"] = raw_response
        return parsed

    def evaluate_versions(
        self,
        poem_versions: Dict[str, Sequence[str] | str | None],
        start_line: str,
    ) -> List[Dict[str, Any]]:
        """Evaluate all poem versions in a single API call."""

        entries = [
            (name, normalize_poem(poem))
            for name, poem in poem_versions.items()
            if normalize_poem(poem)
        ]
        if not entries:
            return []

        poems_block = "\n".join(
            f"Стихотворение {i}:\n{chr(10).join(lines)}"
            for i, (_, lines) in enumerate(entries, 1)
        )
        n = len(entries)

        prompt = f"""
Ты независимый эксперт по оценке русскоязычных стихотворений.
Оцени каждое из {n} стихотворений по шкале от 1 до 5.
1 - очень плохо, 2 - плохо, 3 - средне, 4 - хорошо, 5 - отлично.

Используй всю шкалу 1-5. Оценивай каждый критерий независимо.
Ставь 1 только если текст почти полностью распадается по критерию.

Критерии:
- semantic_coherence: смысловая связность;
- grammar: грамматическая правильность;
- theme_consistency: соответствие теме первой строки;
- poetic_quality: художественность и выразительность;
- overall: общее качество.

Первая строка (тема для всех): "{start_line}"

{poems_block}

Верни JSON-массив из {n} объектов (по порядку стихотворений):
[
  {{"id": 1, "semantic_coherence": <число 1-5>, "grammar": <число 1-5>, "theme_consistency": <число 1-5>, "poetic_quality": <число 1-5>, "overall": <число 1-5>, "comment": "краткий комментарий"}},
  ...
]
Только JSON, без markdown-разметки.
""".strip()

        raw = self._generate(
            prompt,
            temperature=self.config.evaluator_temperature,
            response_mime_type="application/json",
        )

        parsed_list = self._parse_evaluation_batch(raw, n)

        rows = []
        for i, (name, _) in enumerate(entries):
            data = parsed_list[i] if i < len(parsed_list) else {}
            row = {"name": name}
            for field in SCORE_FIELDS:
                row[field] = self._normalize_score(data.get(field))
            row["comment"] = str(data.get("comment", "")).strip()
            rows.append(row)
        return rows

    def _parse_evaluation_batch(
        self, raw_response: str, expected: int
    ) -> List[Dict[str, Any]]:
        text = raw_response.strip()
        text = re.sub(r"^```(?:json)?", "", text).strip()
        text = re.sub(r"```$", "", text).strip()
        text = re.sub(r'"([^"]+)":\s*<[^>]+>', r'"\1": 0', text)

        match = re.search(r"\[.*\]", text, flags=re.S)
        if match:
            text = match.group(0)

        try:
            data = json.loads(text)
            if isinstance(data, list):
                return data
        except Exception:
            pass

        # Fallback: extract individual objects
        objects = re.findall(r"\{[^{}]+\}", text, flags=re.S)
        result = []
        for obj_text in objects:
            try:
                result.append(json.loads(obj_text))
            except Exception:
                obj = {}
                for field in SCORE_FIELDS:
                    m = re.search(rf'"{re.escape(field)}"\s*:\s*(\d+)', obj_text)
                    if m:
                        obj[field] = int(m.group(1))
                result.append(obj)
        return result[:expected] if result else [{}] * expected

    def _parse_evaluation(self, raw_response: str) -> Dict[str, Any]:
        text = raw_response.strip()
        text = re.sub(r"^```(?:json)?", "", text).strip()
        text = re.sub(r"```$", "", text).strip()

        # Replace placeholder values like <число от 1 до 5> with 0 so JSON parses.
        text = re.sub(r'"([^"]+)":\s*<[^>]+>', r'"\1": 0', text)

        match = re.search(r"\{.*\}", text, flags=re.S)
        if match:
            text = match.group(0)

        try:
            data = json.loads(text)
        except Exception:
            # Fallback: extract field:number pairs directly from raw text.
            data = {}
            for field in SCORE_FIELDS:
                m = re.search(
                    rf'"{re.escape(field)}"\s*:\s*(\d+)', raw_response
                )
                if m:
                    data[field] = int(m.group(1))
            comment_m = re.search(r'"comment"\s*:\s*"([^"]*)"', raw_response)
            data["comment"] = comment_m.group(1) if comment_m else ""
            if not any(field in data for field in SCORE_FIELDS):
                return self._empty_evaluation("Не удалось разобрать JSON-оценку.")

        normalized: Dict[str, Any] = {}
        for field in SCORE_FIELDS:
            normalized[field] = self._normalize_score(data.get(field))
        normalized["comment"] = str(data.get("comment", "")).strip()
        return normalized

    @staticmethod
    def _normalize_score(value: Any) -> int:
        try:
            score = int(round(float(value)))
        except Exception:
            score = 0
        return max(1, min(5, score)) if score else 0

    @staticmethod
    def _empty_evaluation(comment: str) -> Dict[str, Any]:
        data = {field: 0 for field in SCORE_FIELDS}
        data["comment"] = comment
        return data
