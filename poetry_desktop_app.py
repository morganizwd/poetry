#!/usr/bin/env python3
"""Desktop GUI for the local poetry generation pipeline."""

from __future__ import annotations

import csv
import json
import os
import queue
import shutil
import subprocess
import sys
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText
from typing import Any, Dict, List, Optional, Sequence

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

from llm_poetry_tools import DEFAULT_OLLAMA_MODEL
from poetry_local_pipeline import (
    DEFAULT_POEM_LINES,
    DEFAULT_RHYME_SCHEME,
    DEFAULT_START_LINE,
)


APP_TITLE = "Poetry Desktop"
DEFAULT_OLLAMA_BASE_URL = "http://127.0.0.1:11434"
PRESET_DEFINITIONS: List[Dict[str, Any]] = [
    {
        "label": "Быстрый старт (рекомендуется)",
        "description": (
            "Для первого знакомства с проектом. Подготовит данные, покажет Markov и "
            "быструю LSTM, но не будет тратить время на долгое обучение."
        ),
        "settings": {
            "fast_mode": True,
            "prepare_data_only": False,
            "enable_llm_evaluation": False,
            "skip_llm_editing": True,
            "skip_quick_lstm": False,
            "skip_full_lstm": True,
            "skip_batch": True,
            "no_lstm_cache": False,
            "force_dataset_refresh": False,
            "quick_epochs": 1,
            "full_epochs": 5,
            "experiment_runs": 1,
            "experiment_llm_runs": 1,
            "poem_lines": 8,
            "llm_provider": "disabled",
        },
    },
    {
        "label": "Обычный локальный запуск",
        "description": (
            "Подходит, если хотите получить полноценный результат без LLM. "
            "Запускает Markov и основную LSTM, но без лишнего batch-сравнения."
        ),
        "settings": {
            "fast_mode": False,
            "prepare_data_only": False,
            "enable_llm_evaluation": False,
            "skip_llm_editing": True,
            "skip_quick_lstm": True,
            "skip_full_lstm": False,
            "skip_batch": True,
            "no_lstm_cache": False,
            "force_dataset_refresh": False,
            "quick_epochs": 1,
            "full_epochs": 6,
            "experiment_runs": 1,
            "experiment_llm_runs": 1,
            "poem_lines": 8,
            "llm_provider": "disabled",
        },
    },
    {
        "label": "Черновик + улучшение через Ollama",
        "description": (
            "Сначала генерирует стих локально, потом просит локальную Ollama его улучшить. "
            "Перед запуском нужен работающий Ollama и скачанная модель."
        ),
        "settings": {
            "fast_mode": True,
            "prepare_data_only": False,
            "enable_llm_evaluation": False,
            "skip_llm_editing": False,
            "skip_quick_lstm": False,
            "skip_full_lstm": True,
            "skip_batch": True,
            "no_lstm_cache": False,
            "force_dataset_refresh": False,
            "quick_epochs": 1,
            "full_epochs": 5,
            "experiment_runs": 1,
            "experiment_llm_runs": 1,
            "poem_lines": 8,
            "llm_provider": "ollama",
        },
    },
    {
        "label": "Только подготовить данные",
        "description": (
            "Скачивает и подготавливает датасет, но не обучает модели и не генерирует стихи."
        ),
        "settings": {
            "fast_mode": True,
            "prepare_data_only": True,
            "enable_llm_evaluation": False,
            "skip_llm_editing": True,
            "skip_quick_lstm": True,
            "skip_full_lstm": True,
            "skip_batch": True,
            "no_lstm_cache": False,
            "force_dataset_refresh": False,
            "quick_epochs": 1,
            "full_epochs": 5,
            "experiment_runs": 1,
            "experiment_llm_runs": 1,
            "poem_lines": 8,
            "llm_provider": "disabled",
        },
    },
]
DEFAULT_PRESET_LABEL = PRESET_DEFINITIONS[0]["label"]
PRESET_BY_LABEL = {preset["label"]: preset for preset in PRESET_DEFINITIONS}


def detect_preferred_python(project_dir: Path) -> Path:
    local_venv_python = project_dir / ".venv" / "Scripts" / "python.exe"
    if local_venv_python.exists():
        return local_venv_python
    return Path(sys.executable).resolve()


def open_path(path: Path) -> None:
    if path.suffix:
        path.parent.mkdir(parents=True, exist_ok=True)
    else:
        path.mkdir(parents=True, exist_ok=True)
    if sys.platform.startswith("win"):
        os.startfile(str(path))
        return
    if sys.platform == "darwin":
        subprocess.Popen(["open", str(path)])
        return
    subprocess.Popen(["xdg-open", str(path)])


def install_entry_shortcuts(widget: tk.Entry | ttk.Entry) -> None:
    """Add explicit clipboard shortcuts and a context menu for text inputs."""

    def is_readonly() -> bool:
        try:
            return str(widget.cget("state")) == "readonly"
        except Exception:
            return False

    def copy_text(_event: object | None = None) -> str:
        try:
            selected_text = widget.selection_get()
        except Exception:
            return "break"
        widget.clipboard_clear()
        widget.clipboard_append(selected_text)
        return "break"

    def cut_text(_event: object | None = None) -> str:
        if is_readonly():
            return "break"
        try:
            selected_text = widget.selection_get()
            first = widget.index("sel.first")
            last = widget.index("sel.last")
        except Exception:
            return "break"
        widget.clipboard_clear()
        widget.clipboard_append(selected_text)
        widget.delete(first, last)
        return "break"

    def paste_text(_event: object | None = None) -> str:
        if is_readonly():
            return "break"
        try:
            clipboard_text = widget.clipboard_get()
        except Exception:
            return "break"

        try:
            first = widget.index("sel.first")
            last = widget.index("sel.last")
            widget.delete(first, last)
            widget.insert(first, clipboard_text)
        except Exception:
            widget.insert(tk.INSERT, clipboard_text)
        return "break"

    def select_all(_event: object | None = None) -> str:
        try:
            widget.selection_range(0, tk.END)
            widget.icursor(tk.END)
        except Exception:
            return "break"
        return "break"

    menu = tk.Menu(widget, tearoff=0)
    menu.add_command(label="Вырезать", command=cut_text)
    menu.add_command(label="Копировать", command=copy_text)
    menu.add_command(label="Вставить", command=paste_text)
    menu.add_separator()
    menu.add_command(label="Выделить всё", command=select_all)

    def show_menu(event: tk.Event[tk.Misc]) -> str:
        try:
            widget.focus_set()
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
        return "break"

    widget.bind("<Control-v>", paste_text, add="+")
    widget.bind("<Control-V>", paste_text, add="+")
    widget.bind("<Shift-Insert>", paste_text, add="+")
    widget.bind("<Control-c>", copy_text, add="+")
    widget.bind("<Control-C>", copy_text, add="+")
    widget.bind("<Control-x>", cut_text, add="+")
    widget.bind("<Control-X>", cut_text, add="+")
    widget.bind("<Control-a>", select_all, add="+")
    widget.bind("<Control-A>", select_all, add="+")
    widget.bind("<Button-3>", show_menu, add="+")


def read_csv_rows(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        return list(csv.DictReader(csv_file))


def read_excel_rows(path: Path) -> List[Dict[str, str]]:
    if not path.exists() or load_workbook is None:
        return []

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        workbook.close()
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    if not any(headers):
        workbook.close()
        return []

    parsed_rows: List[Dict[str, str]] = []
    for row in rows_iter:
        parsed_rows.append(
            {
                headers[index]: "" if value is None else str(value)
                for index, value in enumerate(row[: len(headers)])
                if headers[index]
            }
        )
    workbook.close()
    return parsed_rows


def read_report_rows(path: Path) -> List[Dict[str, str]]:
    if path.suffix.lower() == ".xlsx":
        return read_excel_rows(path)
    return read_csv_rows(path)


def latest_report_files(reports_dir: Path, suffix: str) -> List[Path]:
    xlsx_candidates = (
        list(reports_dir.glob(f"*{suffix}*.xlsx")) if load_workbook is not None else []
    )
    candidates = xlsx_candidates + list(reports_dir.glob(f"*{suffix}*.csv"))
    return sorted(
        candidates,
        key=lambda item: (item.stat().st_mtime, item.suffix.lower() == ".xlsx"),
        reverse=True,
    )


def detect_nvidia_gpu_text() -> str:
    try:
        completed = subprocess.run(
            [
                "nvidia-smi",
                "--query-gpu=name,driver_version",
                "--format=csv,noheader",
            ],
            check=True,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
    except Exception:
        return "NVIDIA GPU: не обнаружен через nvidia-smi"

    lines = [line.strip() for line in completed.stdout.splitlines() if line.strip()]
    if not lines:
        return "NVIDIA GPU: не обнаружен через nvidia-smi"
    return "NVIDIA GPU: " + " | ".join(lines)


def detect_tensorflow_gpu_text(python_executable: Path) -> str:
    code = (
        "import tensorflow as tf; "
        "devices=tf.config.list_physical_devices('GPU'); "
        "print('|'.join(getattr(d,'name',str(d)) for d in devices))"
    )
    try:
        completed = subprocess.run(
            [str(python_executable), "-c", code],
            check=True,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
    except Exception as exc:
        return f"TensorFlow GPU: ошибка проверки ({exc})"

    stdout = completed.stdout.strip()
    if stdout:
        return f"TensorFlow GPU: {stdout}"
    return (
        "TensorFlow GPU: локальный GPU не используется. "
        "На native Windows TensorFlow 2.11+ работает без GPU."
    )


def format_metrics_text(row: Dict[str, str]) -> str:
    parts = [
        f"scope={row.get('report_scope', '')}",
        f"run={row.get('run', '')}",
        f"version={row.get('version', '')}",
        f"stage={row.get('stage', '')}",
        f"source_model={row.get('source_model', '')}",
        f"lines={row.get('actual_lines', row.get('lines', ''))}",
        f"rhyme_success={row.get('rhyme_success', '')}",
        f"rhyme_quality={row.get('rhyme_quality', '')}",
        f"unique_rate={row.get('unique_rate', '')}",
        f"unique_words={row.get('unique_words', '')}",
    ]
    if row.get("overall"):
        parts.append(f"overall={row.get('overall', '')}")
    if row.get("semantic_coherence"):
        parts.append(f"semantic={row.get('semantic_coherence', '')}")
    if row.get("grammar"):
        parts.append(f"grammar={row.get('grammar', '')}")
    return " | ".join(part for part in parts if not part.endswith("="))


class PipelineDesktopApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1380x930")
        self.root.minsize(1120, 760)

        self.project_dir = Path(__file__).resolve().parent
        self.python_executable = detect_preferred_python(self.project_dir)
        self.pipeline_path = self.project_dir / "poetry_local_pipeline.py"
        self.data_dir = self.project_dir / "data"
        self.reports_dir = self.project_dir / "reports"
        self.cache_dir = self.project_dir / "cache" / "poetry_lstm_cache"

        self.process: subprocess.Popen[str] | None = None
        self.reader_thread: threading.Thread | None = None
        self.log_queue: queue.Queue[tuple[str, str]] = queue.Queue()

        self.detailed_rows: List[Dict[str, str]] = []
        self.summary_rows: List[Dict[str, str]] = []
        self.compare_pairs: List[Dict[str, Any]] = []
        self.detailed_reports: List[Path] = []
        self.summary_reports: List[Path] = []

        self._build_variables()
        self._build_layout()
        self._register_variable_traces()
        self._apply_selected_preset(announce=False)
        self._append_log(
            "info",
            "Приложение готово. Оно запускает локальный pipeline без notebook, "
            "показывает лог, загружает отчёты и умеет сравнивать raw/LLM версии.\n"
            f"Python для pipeline: {self.python_executable}\n",
        )
        self._refresh_hardware_async()
        self._refresh_reports()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.bind_all("<MouseWheel>", self._handle_global_mousewheel, add="+")
        self.root.bind_all("<Button-4>", self._handle_global_mousewheel, add="+")
        self.root.bind_all("<Button-5>", self._handle_global_mousewheel, add="+")
        self.root.after(100, self._drain_log_queue)

    def _build_variables(self) -> None:
        self.dataset_txt_var = tk.StringVar()
        self.start_line_var = tk.StringVar(value=DEFAULT_START_LINE)
        self.rhyme_scheme_var = tk.StringVar(value=DEFAULT_RHYME_SCHEME)
        self.poem_lines_var = tk.IntVar(value=DEFAULT_POEM_LINES)
        self.quick_epochs_var = tk.IntVar(value=1)
        self.full_epochs_var = tk.IntVar(value=20)
        self.experiment_runs_var = tk.IntVar(value=2)
        self.experiment_llm_runs_var = tk.IntVar(value=2)
        self.llm_provider_var = tk.StringVar(value="disabled")
        self.llm_model_name_var = tk.StringVar(value=DEFAULT_OLLAMA_MODEL)
        self.ollama_base_url_var = tk.StringVar(
            value=os.getenv("OLLAMA_BASE_URL", DEFAULT_OLLAMA_BASE_URL)
        )
        self.preset_var = tk.StringVar(value=DEFAULT_PRESET_LABEL)
        self.preset_description_var = tk.StringVar(value="")
        self.run_summary_var = tk.StringVar(value="")

        self.fast_mode_var = tk.BooleanVar(value=True)
        self.prepare_data_only_var = tk.BooleanVar(value=False)
        self.enable_llm_evaluation_var = tk.BooleanVar(value=False)
        self.skip_llm_editing_var = tk.BooleanVar(value=False)
        self.skip_quick_lstm_var = tk.BooleanVar(value=False)
        self.skip_full_lstm_var = tk.BooleanVar(value=False)
        self.skip_batch_var = tk.BooleanVar(value=False)
        self.no_lstm_cache_var = tk.BooleanVar(value=False)
        self.force_dataset_refresh_var = tk.BooleanVar(value=False)

        self.status_var = tk.StringVar(
            value="Готово к запуску. Если не уверены, начните с пресета 'Быстрый старт'."
        )
        self.llm_status_var = tk.StringVar(
            value=(
                "Ollama необязательна. Если хотите LLM-редактирование, включите режим "
                "'ollama', запустите локальный сервер и нажмите 'Проверить Ollama'."
            )
        )
        self.nvidia_status_var = tk.StringVar(value="NVIDIA GPU: проверяем...")
        self.tensorflow_gpu_status_var = tk.StringVar(value="TensorFlow GPU: проверяем...")

        self.detailed_report_var = tk.StringVar()
        self.summary_report_var = tk.StringVar()
        self.report_info_var = tk.StringVar(
            value="После первого запуска здесь появятся подробные отчёты Excel/CSV."
        )

    def _build_layout(self) -> None:
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        header = ttk.Frame(self.root, padding=(16, 16, 16, 8))
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)

        ttk.Label(
            header,
            text="Poetry Desktop",
            font=("Segoe UI", 18, "bold"),
        ).grid(row=0, column=0, sticky="w")
        ttk.Label(
            header,
            text=(
                "Обычный desktop-вариант проекта: запуск пайплайна, Ollama-настройка, "
                "просмотр всех сгенерированных стихов и сравнение raw vs LLM."
            ),
            wraplength=1100,
            justify="left",
        ).grid(row=1, column=0, sticky="w", pady=(6, 0))

        notebook = ttk.Notebook(self.root)
        notebook.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 16))

        run_tab = ttk.Frame(notebook, padding=8)
        poems_tab = ttk.Frame(notebook, padding=8)
        compare_tab = ttk.Frame(notebook, padding=8)
        reports_tab = ttk.Frame(notebook, padding=8)

        notebook.add(run_tab, text="Запуск")
        notebook.add(poems_tab, text="Стихи")
        notebook.add(compare_tab, text="Сравнение")
        notebook.add(reports_tab, text="Отчёты")

        self._build_run_tab_friendly(run_tab)
        self._build_poems_tab(poems_tab)
        self._build_compare_tab(compare_tab)
        self._build_reports_tab(reports_tab)

    def _build_run_tab(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(0, weight=1)

        controls = ttk.Frame(parent)
        controls.grid(row=0, column=0, sticky="nsw", padx=(0, 10))
        controls.columnconfigure(0, weight=1)

        log_frame = ttk.Frame(parent)
        log_frame.grid(row=0, column=1, sticky="nsew")
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(1, weight=1)

        row = 0

        run_box = ttk.LabelFrame(controls, text="Запуск", padding=10)
        run_box.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        run_box.columnconfigure(0, weight=1)
        run_box.columnconfigure(1, weight=1)
        run_box.columnconfigure(2, weight=1)

        ttk.Button(run_box, text="Быстрый запуск", command=self._run_fast_mode).grid(
            row=0, column=0, sticky="ew", padx=(0, 6)
        )
        ttk.Button(run_box, text="Полный запуск", command=self._run_full_mode).grid(
            row=0, column=1, sticky="ew", padx=(0, 6)
        )
        ttk.Button(run_box, text="Только данные", command=self._run_prepare_only).grid(
            row=0, column=2, sticky="ew"
        )

        ttk.Button(run_box, text="Остановить", command=self._stop_process).grid(
            row=1, column=0, sticky="ew", padx=(0, 6), pady=(8, 0)
        )
        ttk.Button(run_box, text="Очистить лог", command=self._clear_log).grid(
            row=1, column=1, sticky="ew", padx=(0, 6), pady=(8, 0)
        )
        ttk.Button(
            run_box,
            text="Скопировать команду",
            command=self._copy_command_preview,
        ).grid(row=1, column=2, sticky="ew", pady=(8, 0))
        ttk.Button(
            run_box,
            text="Удалить веса LSTM",
            command=self._delete_lstm_weights,
        ).grid(row=2, column=0, columnspan=3, sticky="ew", pady=(8, 0))

        row += 1

        llm_box = ttk.LabelFrame(controls, text="LLM / Ollama", padding=10)
        llm_box.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        llm_box.columnconfigure(1, weight=1)

        ttk.Label(llm_box, text="Провайдер").grid(row=0, column=0, sticky="w", pady=(0, 6))
        ttk.Combobox(
            llm_box,
            textvariable=self.llm_provider_var,
            values=["ollama", "disabled"],
            state="readonly",
        ).grid(row=0, column=1, sticky="ew", pady=(0, 6))

        ttk.Label(llm_box, text="Ollama model").grid(row=1, column=0, sticky="w", pady=6)
        llm_model_entry = ttk.Entry(llm_box, textvariable=self.llm_model_name_var)
        llm_model_entry.grid(row=1, column=1, sticky="ew", pady=6)
        install_entry_shortcuts(llm_model_entry)

        ttk.Label(llm_box, text="Ollama URL").grid(row=2, column=0, sticky="w", pady=6)
        ollama_url_entry = ttk.Entry(llm_box, textvariable=self.ollama_base_url_var)
        ollama_url_entry.grid(row=2, column=1, sticky="ew", pady=6)
        install_entry_shortcuts(ollama_url_entry)

        ttk.Button(llm_box, text="Проверить Ollama", command=self._check_ollama_async).grid(
            row=3, column=0, sticky="ew", pady=(8, 0), padx=(0, 6)
        )
        ttk.Button(
            llm_box,
            text="URL по умолчанию",
            command=lambda: self.ollama_base_url_var.set(DEFAULT_OLLAMA_BASE_URL),
        ).grid(row=3, column=1, sticky="w", pady=(8, 0))

        ttk.Label(
            llm_box,
            textvariable=self.llm_status_var,
            wraplength=360,
            justify="left",
        ).grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 0))

        row += 1

        hardware_box = ttk.LabelFrame(controls, text="GPU / железо", padding=10)
        hardware_box.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        hardware_box.columnconfigure(0, weight=1)

        ttk.Button(
            hardware_box,
            text="Обновить статус GPU",
            command=self._refresh_hardware_async,
        ).grid(row=0, column=0, sticky="ew", pady=(0, 8))
        ttk.Label(
            hardware_box,
            textvariable=self.nvidia_status_var,
            wraplength=360,
            justify="left",
        ).grid(row=1, column=0, sticky="w", pady=2)
        ttk.Label(
            hardware_box,
            textvariable=self.tensorflow_gpu_status_var,
            wraplength=360,
            justify="left",
        ).grid(row=2, column=0, sticky="w", pady=2)
        ttk.Label(
            hardware_box,
            text=(
                "Важно: на native Windows TensorFlow 2.11+ локально не использует GPU. "
                "Для LSTM это ограничение остаётся, но локальная Ollama работает отдельно "
                "от TensorFlow и может использовать свой runtime."
            ),
            wraplength=360,
            justify="left",
        ).grid(row=3, column=0, sticky="w", pady=(8, 0))

        row += 1

        main_box = ttk.LabelFrame(controls, text="Основные настройки", padding=10)
        main_box.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        main_box.columnconfigure(1, weight=1)

        ttk.Label(main_box, text="Стартовая строка").grid(row=0, column=0, sticky="w", pady=(0, 6))
        start_line_entry = ttk.Entry(main_box, textvariable=self.start_line_var)
        start_line_entry.grid(row=0, column=1, sticky="ew", pady=(0, 6))
        install_entry_shortcuts(start_line_entry)

        ttk.Label(main_box, text="Схема рифмы").grid(row=1, column=0, sticky="w", pady=6)
        ttk.Combobox(
            main_box,
            textvariable=self.rhyme_scheme_var,
            values=["AABB", "ABAB", "ABBA", "AAAA"],
            state="readonly",
        ).grid(row=1, column=1, sticky="ew", pady=6)

        ttk.Label(main_box, text="Количество строк").grid(row=2, column=0, sticky="w", pady=6)
        ttk.Spinbox(
            main_box,
            from_=2,
            to=16,
            textvariable=self.poem_lines_var,
            width=8,
        ).grid(row=2, column=1, sticky="w", pady=6)

        ttk.Label(main_box, text="Ваш poems_clean.txt").grid(row=3, column=0, sticky="w", pady=(6, 0))
        dataset_row = ttk.Frame(main_box)
        dataset_row.grid(row=3, column=1, sticky="ew", pady=(6, 0))
        dataset_row.columnconfigure(0, weight=1)
        dataset_entry = ttk.Entry(dataset_row, textvariable=self.dataset_txt_var)
        dataset_entry.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        install_entry_shortcuts(dataset_entry)
        ttk.Button(dataset_row, text="Выбрать", command=self._pick_dataset).grid(
            row=0, column=1, sticky="ew"
        )

        row += 1

        train_box = ttk.LabelFrame(controls, text="Обучение и эксперимент", padding=10)
        train_box.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        train_box.columnconfigure(1, weight=1)
        self._add_spinbox_row(train_box, 0, "Quick LSTM epochs", self.quick_epochs_var, 1, 10)
        self._add_spinbox_row(train_box, 1, "Full LSTM epochs", self.full_epochs_var, 1, 100)
        self._add_spinbox_row(train_box, 2, "Batch runs", self.experiment_runs_var, 1, 100)
        self._add_spinbox_row(
            train_box,
            3,
            "Batch LLM runs",
            self.experiment_llm_runs_var,
            1,
            100,
        )

        row += 1

        flags_box = ttk.LabelFrame(controls, text="Флаги запуска", padding=10)
        flags_box.grid(row=row, column=0, sticky="ew")
        flags_box.columnconfigure(0, weight=1)
        flags_box.columnconfigure(1, weight=1)

        flag_widgets = [
            ("Fast mode", self.fast_mode_var, 0, 0),
            ("Prepare data only", self.prepare_data_only_var, 0, 1),
            ("Enable LLM evaluation", self.enable_llm_evaluation_var, 1, 0),
            ("Skip LLM editing", self.skip_llm_editing_var, 1, 1),
            ("Skip quick LSTM", self.skip_quick_lstm_var, 2, 0),
            ("Skip full LSTM", self.skip_full_lstm_var, 2, 1),
            ("Skip batch", self.skip_batch_var, 3, 0),
            ("Disable LSTM cache", self.no_lstm_cache_var, 3, 1),
            ("Force dataset refresh", self.force_dataset_refresh_var, 4, 0),
        ]
        for text, variable, row_index, col_index in flag_widgets:
            ttk.Checkbutton(flags_box, text=text, variable=variable).grid(
                row=row_index,
                column=col_index,
                sticky="w",
                pady=2,
                padx=(0, 12),
            )

        ttk.Label(
            log_frame,
            text="Лог выполнения пайплайна",
            font=("Segoe UI", 12, "bold"),
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        self.log_widget = ScrolledText(
            log_frame,
            wrap=tk.WORD,
            font=("Consolas", 10),
            state="disabled",
        )
        self.log_widget.grid(row=1, column=0, sticky="nsew")

        ttk.Label(
            log_frame,
            textvariable=self.status_var,
            wraplength=820,
            justify="left",
        ).grid(row=2, column=0, sticky="w", pady=(8, 0))

    def _build_run_tab_friendly(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(0, weight=1)

        controls_host = ttk.Frame(parent)
        controls_host.grid(row=0, column=0, sticky="ns", padx=(0, 10))
        controls_host.columnconfigure(0, weight=1)
        controls_host.rowconfigure(0, weight=1)

        self.run_controls_canvas = tk.Canvas(
            controls_host,
            highlightthickness=0,
            borderwidth=0,
            width=430,
        )
        self.run_controls_canvas.grid(row=0, column=0, sticky="nsew")

        self.run_controls_scrollbar = ttk.Scrollbar(
            controls_host,
            orient="vertical",
            command=self.run_controls_canvas.yview,
        )
        self.run_controls_scrollbar.grid(row=0, column=1, sticky="ns")
        self.run_controls_canvas.configure(
            yscrollcommand=self.run_controls_scrollbar.set
        )

        controls = ttk.Frame(self.run_controls_canvas)
        controls.columnconfigure(0, weight=1)
        self.run_controls_window = self.run_controls_canvas.create_window(
            (0, 0),
            window=controls,
            anchor="nw",
        )
        controls.bind("<Configure>", self._sync_run_controls_scrollregion)
        self.run_controls_canvas.bind("<Configure>", self._sync_run_controls_canvas_width)
        self.root.after_idle(self._sync_run_controls_scrollregion)

        log_frame = ttk.Frame(parent)
        log_frame.grid(row=0, column=1, sticky="nsew")
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(1, weight=1)

        row = 0

        guide_box = ttk.LabelFrame(controls, text="С чего начать", padding=10)
        guide_box.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        guide_box.columnconfigure(0, weight=1)
        ttk.Label(
            guide_box,
            text=(
                "1. Выберите готовый сценарий.\n"
                "2. Если нужен LLM, включите Ollama и нажмите 'Проверить Ollama'.\n"
                "3. Нажмите 'Запустить по текущим настройкам'.\n"
                "После завершения стихи и отчёты автоматически появятся во вкладках."
            ),
            wraplength=360,
            justify="left",
        ).grid(row=0, column=0, sticky="w")

        row += 1

        preset_box = ttk.LabelFrame(controls, text="Готовые сценарии", padding=10)
        preset_box.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        preset_box.columnconfigure(1, weight=1)

        ttk.Label(preset_box, text="Сценарий").grid(row=0, column=0, sticky="w", pady=(0, 6))
        self.preset_combo = ttk.Combobox(
            preset_box,
            textvariable=self.preset_var,
            values=[preset["label"] for preset in PRESET_DEFINITIONS],
            state="readonly",
        )
        self.preset_combo.grid(row=0, column=1, sticky="ew", pady=(0, 6))
        self.preset_combo.bind(
            "<<ComboboxSelected>>",
            lambda _event: self._apply_selected_preset(),
        )

        ttk.Button(preset_box, text="Применить", command=self._apply_selected_preset).grid(
            row=1, column=0, sticky="ew", pady=(4, 0), padx=(0, 6)
        )
        ttk.Button(
            preset_box,
            text="Применить и запустить",
            command=self._run_selected_preset,
        ).grid(row=1, column=1, sticky="ew", pady=(4, 0))

        ttk.Label(
            preset_box,
            textvariable=self.preset_description_var,
            wraplength=360,
            justify="left",
        ).grid(row=2, column=0, columnspan=2, sticky="w", pady=(8, 0))
        ttk.Label(
            preset_box,
            textvariable=self.run_summary_var,
            wraplength=360,
            justify="left",
        ).grid(row=3, column=0, columnspan=2, sticky="w", pady=(8, 0))

        row += 1

        run_box = ttk.LabelFrame(controls, text="Действия", padding=10)
        run_box.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        run_box.columnconfigure(0, weight=1)
        run_box.columnconfigure(1, weight=1)
        run_box.columnconfigure(2, weight=1)

        ttk.Button(
            run_box,
            text="Запустить по текущим настройкам",
            command=self._start_process,
        ).grid(row=0, column=0, columnspan=2, sticky="ew", padx=(0, 6))
        ttk.Button(run_box, text="Остановить", command=self._stop_process).grid(
            row=0, column=2, sticky="ew"
        )
        ttk.Button(run_box, text="Очистить лог", command=self._clear_log).grid(
            row=1, column=0, sticky="ew", padx=(0, 6), pady=(8, 0)
        )
        ttk.Button(
            run_box,
            text="Скопировать команду",
            command=self._copy_command_preview,
        ).grid(row=1, column=1, sticky="ew", padx=(0, 6), pady=(8, 0))
        ttk.Button(
            run_box,
            text="Удалить веса LSTM",
            command=self._delete_lstm_weights,
        ).grid(row=1, column=2, sticky="ew", pady=(8, 0))

        row += 1

        llm_box = ttk.LabelFrame(controls, text="Локальная Ollama", padding=10)
        llm_box.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        llm_box.columnconfigure(1, weight=1)

        ttk.Label(llm_box, text="Режим LLM").grid(row=0, column=0, sticky="w", pady=(0, 6))
        ttk.Combobox(
            llm_box,
            textvariable=self.llm_provider_var,
            values=["disabled", "ollama"],
            state="readonly",
        ).grid(row=0, column=1, sticky="ew", pady=(0, 6))

        ttk.Label(llm_box, text="Модель Ollama").grid(row=1, column=0, sticky="w", pady=6)
        ollama_model_entry = ttk.Entry(llm_box, textvariable=self.llm_model_name_var)
        ollama_model_entry.grid(row=1, column=1, sticky="ew", pady=6)
        install_entry_shortcuts(ollama_model_entry)

        ttk.Label(llm_box, text="Адрес Ollama").grid(row=2, column=0, sticky="w", pady=6)
        ollama_url_entry = ttk.Entry(llm_box, textvariable=self.ollama_base_url_var)
        ollama_url_entry.grid(row=2, column=1, sticky="ew", pady=6)
        install_entry_shortcuts(ollama_url_entry)

        ttk.Button(llm_box, text="Проверить Ollama", command=self._check_ollama_async).grid(
            row=3, column=0, sticky="ew", pady=(8, 0), padx=(0, 6)
        )
        ttk.Button(
            llm_box,
            text="URL по умолчанию",
            command=lambda: self.ollama_base_url_var.set(DEFAULT_OLLAMA_BASE_URL),
        ).grid(row=3, column=1, sticky="w", pady=(8, 0))

        ttk.Label(
            llm_box,
            textvariable=self.llm_status_var,
            wraplength=360,
            justify="left",
        ).grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 0))

        row += 1

        hardware_box = ttk.LabelFrame(controls, text="GPU и железо", padding=10)
        hardware_box.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        hardware_box.columnconfigure(0, weight=1)

        ttk.Button(
            hardware_box,
            text="Обновить статус GPU",
            command=self._refresh_hardware_async,
        ).grid(row=0, column=0, sticky="ew", pady=(0, 8))
        ttk.Label(
            hardware_box,
            textvariable=self.nvidia_status_var,
            wraplength=360,
            justify="left",
        ).grid(row=1, column=0, sticky="w", pady=2)
        ttk.Label(
            hardware_box,
            textvariable=self.tensorflow_gpu_status_var,
            wraplength=360,
            justify="left",
        ).grid(row=2, column=0, sticky="w", pady=2)
        ttk.Label(
            hardware_box,
            text=(
                "Важно: на native Windows TensorFlow 2.11+ не использует GPU для LSTM. "
                "Это ограничение платформы, а не ошибка вашего проекта."
            ),
            wraplength=360,
            justify="left",
        ).grid(row=3, column=0, sticky="w", pady=(8, 0))

        row += 1

        main_box = ttk.LabelFrame(controls, text="Что генерировать", padding=10)
        main_box.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        main_box.columnconfigure(1, weight=1)
        ttk.Label(
            main_box,
            text="Если не менять поля ниже, проект уже готов к первому запуску.",
            wraplength=360,
            justify="left",
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

        ttk.Label(main_box, text="Первая строка").grid(row=1, column=0, sticky="w", pady=(0, 6))
        start_line_entry = ttk.Entry(main_box, textvariable=self.start_line_var)
        start_line_entry.grid(row=1, column=1, sticky="ew", pady=(0, 6))
        install_entry_shortcuts(start_line_entry)

        ttk.Label(main_box, text="Схема рифмы").grid(row=2, column=0, sticky="w", pady=6)
        ttk.Combobox(
            main_box,
            textvariable=self.rhyme_scheme_var,
            values=["AABB", "ABAB", "ABBA", "AAAA"],
            state="readonly",
        ).grid(row=2, column=1, sticky="ew", pady=6)

        ttk.Label(main_box, text="Сколько строк в стихе").grid(row=3, column=0, sticky="w", pady=6)
        ttk.Spinbox(
            main_box,
            from_=2,
            to=16,
            textvariable=self.poem_lines_var,
            width=8,
        ).grid(row=3, column=1, sticky="w", pady=6)

        ttk.Label(main_box, text="Свой poems_clean.txt").grid(row=4, column=0, sticky="w", pady=(6, 0))
        dataset_row = ttk.Frame(main_box)
        dataset_row.grid(row=4, column=1, sticky="ew", pady=(6, 0))
        dataset_row.columnconfigure(0, weight=1)
        dataset_entry = ttk.Entry(dataset_row, textvariable=self.dataset_txt_var)
        dataset_entry.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        install_entry_shortcuts(dataset_entry)
        ttk.Button(dataset_row, text="Выбрать", command=self._pick_dataset).grid(
            row=0, column=1, sticky="ew"
        )

        row += 1

        train_box = ttk.LabelFrame(controls, text="Насколько долго обучать модели", padding=10)
        train_box.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        train_box.columnconfigure(1, weight=1)
        self._add_spinbox_row(train_box, 0, "Эпох быстрой LSTM", self.quick_epochs_var, 1, 10)
        self._add_spinbox_row(train_box, 1, "Эпох основной LSTM", self.full_epochs_var, 1, 100)
        self._add_spinbox_row(train_box, 2, "Сколько batch-запусков", self.experiment_runs_var, 1, 100)
        self._add_spinbox_row(
            train_box,
            3,
            "Сколько batch-запусков отдать LLM",
            self.experiment_llm_runs_var,
            1,
            100,
        )

        row += 1

        flags_box = ttk.LabelFrame(controls, text="Дополнительные флажки", padding=10)
        flags_box.grid(row=row, column=0, sticky="ew")
        flags_box.columnconfigure(0, weight=1)
        flags_box.columnconfigure(1, weight=1)

        flag_widgets = [
            ("Быстрый режим", self.fast_mode_var, 0, 0),
            ("Только подготовить данные", self.prepare_data_only_var, 0, 1),
            ("Включить LLM-оценку", self.enable_llm_evaluation_var, 1, 0),
            ("Не редактировать стихи через LLM", self.skip_llm_editing_var, 1, 1),
            ("Пропустить quick LSTM", self.skip_quick_lstm_var, 2, 0),
            ("Пропустить full LSTM", self.skip_full_lstm_var, 2, 1),
            ("Пропустить batch-сравнение", self.skip_batch_var, 3, 0),
            ("Не использовать кэш LSTM", self.no_lstm_cache_var, 3, 1),
            ("Скачать датасет заново", self.force_dataset_refresh_var, 4, 0),
        ]
        for text, variable, row_index, col_index in flag_widgets:
            ttk.Checkbutton(flags_box, text=text, variable=variable).grid(
                row=row_index,
                column=col_index,
                sticky="w",
                pady=2,
                padx=(0, 12),
            )

        ttk.Label(
            log_frame,
            text="Лог выполнения пайплайна",
            font=("Segoe UI", 12, "bold"),
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        self.log_widget = ScrolledText(
            log_frame,
            wrap=tk.WORD,
            font=("Consolas", 10),
            state="disabled",
        )
        self.log_widget.grid(row=1, column=0, sticky="nsew")

        ttk.Label(
            log_frame,
            textvariable=self.status_var,
            wraplength=820,
            justify="left",
        ).grid(row=2, column=0, sticky="w", pady=(8, 0))

    def _sync_run_controls_scrollregion(self, _event: object | None = None) -> None:
        if not hasattr(self, "run_controls_canvas"):
            return
        bbox = self.run_controls_canvas.bbox("all")
        if bbox:
            self.run_controls_canvas.configure(scrollregion=bbox)

    def _sync_run_controls_canvas_width(self, event: tk.Event[tk.Misc]) -> None:
        if not hasattr(self, "run_controls_canvas") or not hasattr(
            self,
            "run_controls_window",
        ):
            return
        self.run_controls_canvas.itemconfigure(self.run_controls_window, width=event.width)

    def _is_widget_inside_run_controls(self, widget: tk.Misc | None) -> bool:
        canvas = getattr(self, "run_controls_canvas", None)
        while widget is not None:
            if widget == canvas:
                return True
            widget = getattr(widget, "master", None)
        return False

    def _handle_global_mousewheel(self, event: tk.Event[tk.Misc]) -> str | None:
        canvas = getattr(self, "run_controls_canvas", None)
        if canvas is None or not canvas.winfo_exists():
            return None

        pointer_widget = self.root.winfo_containing(event.x_root, event.y_root)
        if not self._is_widget_inside_run_controls(pointer_widget):
            return None

        delta = 0
        if getattr(event, "delta", 0):
            delta = -1 if event.delta > 0 else 1
        elif getattr(event, "num", None) == 4:
            delta = -1
        elif getattr(event, "num", None) == 5:
            delta = 1

        if delta:
            canvas.yview_scroll(delta, "units")
            return "break"
        return None

    def _build_poems_tab(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)
        parent.rowconfigure(2, weight=1)

        toolbar = ttk.Frame(parent)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(1, weight=1)

        ttk.Label(toolbar, text="Подробный отчёт").grid(row=0, column=0, sticky="w")
        self.detailed_report_combo = ttk.Combobox(
            toolbar,
            textvariable=self.detailed_report_var,
            state="readonly",
        )
        self.detailed_report_combo.grid(row=0, column=1, sticky="ew", padx=(8, 8))
        self.detailed_report_combo.bind("<<ComboboxSelected>>", lambda _event: self._load_selected_reports())

        ttk.Button(toolbar, text="Обновить", command=self._refresh_reports).grid(
            row=0, column=2, sticky="ew", padx=(0, 6)
        )
        ttk.Button(toolbar, text="Открыть report", command=self._open_selected_detailed).grid(
            row=0, column=3, sticky="ew", padx=(0, 6)
        )
        ttk.Button(toolbar, text="Сохранить копию", command=self._save_selected_detailed_copy).grid(
            row=0, column=4, sticky="ew"
        )

        tree_frame = ttk.Frame(parent)
        tree_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 8))
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        columns = (
            "report_scope",
            "run",
            "source_model",
            "version",
            "stage",
            "actual_lines",
            "rhyme_quality",
            "overall",
        )
        self.poems_tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            height=14,
        )
        headings = {
            "report_scope": "scope",
            "run": "run",
            "source_model": "source",
            "version": "version",
            "stage": "stage",
            "actual_lines": "lines",
            "rhyme_quality": "rhyme",
            "overall": "overall",
        }
        widths = {
            "report_scope": 90,
            "run": 60,
            "source_model": 100,
            "version": 180,
            "stage": 110,
            "actual_lines": 70,
            "rhyme_quality": 70,
            "overall": 70,
        }
        for column in columns:
            self.poems_tree.heading(column, text=headings[column])
            self.poems_tree.column(column, width=widths[column], anchor="w")
        self.poems_tree.grid(row=0, column=0, sticky="nsew")
        self.poems_tree.bind("<<TreeviewSelect>>", self._on_poem_selected)

        poems_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.poems_tree.yview)
        poems_scrollbar.grid(row=0, column=1, sticky="ns")
        self.poems_tree.configure(yscrollcommand=poems_scrollbar.set)

        details_frame = ttk.LabelFrame(parent, text="Текст и метрики", padding=8)
        details_frame.grid(row=2, column=0, sticky="nsew")
        details_frame.columnconfigure(0, weight=1)
        details_frame.rowconfigure(1, weight=1)

        self.poem_details_label = ttk.Label(
            details_frame,
            text="Выберите стих из таблицы выше. Здесь появится его текст и метрики.",
            wraplength=1240,
            justify="left",
        )
        self.poem_details_label.grid(row=0, column=0, sticky="w", pady=(0, 8))

        self.poem_text_widget = ScrolledText(
            details_frame,
            wrap=tk.WORD,
            font=("Consolas", 10),
            state="disabled",
            height=14,
        )
        self.poem_text_widget.grid(row=1, column=0, sticky="nsew")

    def _build_compare_tab(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(0, weight=1)

        left = ttk.Frame(parent)
        left.grid(row=0, column=0, sticky="nsw", padx=(0, 10))
        left.columnconfigure(0, weight=1)
        left.rowconfigure(1, weight=1)

        right = ttk.Frame(parent)
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.columnconfigure(1, weight=1)
        right.rowconfigure(1, weight=1)

        ttk.Label(
            left,
            text="Сравнение до/после LLM",
            font=("Segoe UI", 12, "bold"),
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        compare_columns = ("report_scope", "run", "source_model")
        self.compare_tree = ttk.Treeview(
            left,
            columns=compare_columns,
            show="headings",
            height=22,
        )
        for column, heading, width in [
            ("report_scope", "scope", 90),
            ("run", "run", 70),
            ("source_model", "model", 120),
        ]:
            self.compare_tree.heading(column, text=heading)
            self.compare_tree.column(column, width=width, anchor="w")
        self.compare_tree.grid(row=1, column=0, sticky="nsew")
        self.compare_tree.bind("<<TreeviewSelect>>", self._on_compare_selected)

        raw_box = ttk.LabelFrame(right, text="До LLM", padding=8)
        raw_box.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=(0, 6))
        raw_box.columnconfigure(0, weight=1)
        raw_box.rowconfigure(1, weight=1)

        self.compare_raw_label = ttk.Label(
            raw_box,
            text="Выберите слева пару raw/LLM.",
            wraplength=520,
            justify="left",
        )
        self.compare_raw_label.grid(row=0, column=0, sticky="w", pady=(0, 8))
        self.compare_raw_text = ScrolledText(
            raw_box,
            wrap=tk.WORD,
            font=("Consolas", 10),
            state="disabled",
        )
        self.compare_raw_text.grid(row=1, column=0, sticky="nsew")

        llm_box = ttk.LabelFrame(right, text="После LLM", padding=8)
        llm_box.grid(row=0, column=1, rowspan=2, sticky="nsew")
        llm_box.columnconfigure(0, weight=1)
        llm_box.rowconfigure(1, weight=1)

        self.compare_llm_label = ttk.Label(
            llm_box,
            text="Выберите слева пару raw/LLM.",
            wraplength=520,
            justify="left",
        )
        self.compare_llm_label.grid(row=0, column=0, sticky="w", pady=(0, 8))
        self.compare_llm_text = ScrolledText(
            llm_box,
            wrap=tk.WORD,
            font=("Consolas", 10),
            state="disabled",
        )
        self.compare_llm_text.grid(row=1, column=0, sticky="nsew")

    def _build_reports_tab(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(2, weight=1)

        top = ttk.Frame(parent)
        top.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        top.columnconfigure(1, weight=1)
        top.columnconfigure(4, weight=1)

        ttk.Label(top, text="Подробный CSV").grid(row=0, column=0, sticky="w")
        self.detailed_reports_combo_reports_tab = ttk.Combobox(
            top,
            textvariable=self.detailed_report_var,
            state="readonly",
        )
        self.detailed_reports_combo_reports_tab.grid(
            row=0, column=1, sticky="ew", padx=(8, 8)
        )
        self.detailed_reports_combo_reports_tab.bind(
            "<<ComboboxSelected>>",
            lambda _event: self._load_selected_reports(),
        )

        ttk.Label(top, text="Краткий CSV").grid(row=0, column=2, sticky="w")
        self.summary_report_combo = ttk.Combobox(
            top,
            textvariable=self.summary_report_var,
            state="readonly",
        )
        self.summary_report_combo.grid(row=0, column=3, sticky="ew", padx=(8, 8))
        self.summary_report_combo.bind(
            "<<ComboboxSelected>>",
            lambda _event: self._load_selected_reports(),
        )

        ttk.Button(top, text="Обновить", command=self._refresh_reports).grid(
            row=0, column=4, sticky="e"
        )

        buttons = ttk.Frame(parent)
        buttons.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        for index in range(4):
            buttons.columnconfigure(index, weight=1)

        ttk.Button(
            buttons,
            text="Открыть detailed",
            command=self._open_selected_detailed,
        ).grid(row=0, column=0, sticky="ew", padx=(0, 6))
        ttk.Button(
            buttons,
            text="Сохранить detailed как",
            command=self._save_selected_detailed_copy,
        ).grid(row=0, column=1, sticky="ew", padx=(0, 6))
        ttk.Button(
            buttons,
            text="Открыть summary",
            command=self._open_selected_summary,
        ).grid(row=0, column=2, sticky="ew", padx=(0, 6))
        ttk.Button(
            buttons,
            text="Сохранить summary как",
            command=self._save_selected_summary_copy,
        ).grid(row=0, column=3, sticky="ew")

        ttk.Label(
            parent,
            textvariable=self.report_info_var,
            wraplength=1240,
            justify="left",
        ).grid(row=2, column=0, sticky="nw", pady=(0, 8))

        summary_frame = ttk.Frame(parent)
        summary_frame.grid(row=3, column=0, sticky="nsew")
        summary_frame.columnconfigure(0, weight=1)
        summary_frame.rowconfigure(0, weight=1)
        parent.rowconfigure(3, weight=1)

        columns = (
            "report_scope",
            "version",
            "stage",
            "row_count",
            "source_model",
            "mean_rhyme_quality",
            "mean_overall",
        )
        self.summary_tree = ttk.Treeview(
            summary_frame,
            columns=columns,
            show="headings",
            height=18,
        )
        for column, heading, width in [
            ("report_scope", "scope", 90),
            ("version", "version", 200),
            ("stage", "stage", 110),
            ("row_count", "rows", 70),
            ("source_model", "source", 100),
            ("mean_rhyme_quality", "mean rhyme", 95),
            ("mean_overall", "mean overall", 95),
        ]:
            self.summary_tree.heading(column, text=heading)
            self.summary_tree.column(column, width=width, anchor="w")
        self.summary_tree.grid(row=0, column=0, sticky="nsew")

        summary_scroll = ttk.Scrollbar(summary_frame, orient="vertical", command=self.summary_tree.yview)
        summary_scroll.grid(row=0, column=1, sticky="ns")
        self.summary_tree.configure(yscrollcommand=summary_scroll.set)

    def _add_spinbox_row(
        self,
        parent: ttk.LabelFrame,
        row: int,
        label: str,
        variable: tk.IntVar,
        from_value: int,
        to_value: int,
    ) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=6)
        ttk.Spinbox(
            parent,
            from_=from_value,
            to=to_value,
            textvariable=variable,
            width=8,
        ).grid(row=row, column=1, sticky="w", pady=6)

    def _safe_int(
        self,
        variable: tk.IntVar,
        fallback: int,
        minimum: Optional[int] = None,
        maximum: Optional[int] = None,
    ) -> int:
        try:
            value = int(variable.get())
        except Exception:
            value = fallback
        if minimum is not None:
            value = max(minimum, value)
        if maximum is not None:
            value = min(maximum, value)
        return value

    def _register_variable_traces(self) -> None:
        watched_variables: List[tk.Variable] = [
            self.preset_var,
            self.start_line_var,
            self.rhyme_scheme_var,
            self.poem_lines_var,
            self.quick_epochs_var,
            self.full_epochs_var,
            self.experiment_runs_var,
            self.experiment_llm_runs_var,
            self.llm_provider_var,
            self.llm_model_name_var,
            self.fast_mode_var,
            self.prepare_data_only_var,
            self.enable_llm_evaluation_var,
            self.skip_llm_editing_var,
            self.skip_quick_lstm_var,
            self.skip_full_lstm_var,
            self.skip_batch_var,
            self.no_lstm_cache_var,
            self.force_dataset_refresh_var,
        ]
        for variable in watched_variables:
            variable.trace_add("write", self._handle_variable_change)

    def _handle_variable_change(self, *_args: object) -> None:
        self._refresh_preset_description()
        self._refresh_run_summary()

    def _selected_preset(self) -> Dict[str, Any]:
        return PRESET_BY_LABEL.get(self.preset_var.get(), PRESET_DEFINITIONS[0])

    def _apply_selected_preset(self, announce: bool = True) -> None:
        preset = self._selected_preset()
        settings = preset.get("settings", {})

        self.fast_mode_var.set(bool(settings.get("fast_mode", False)))
        self.prepare_data_only_var.set(bool(settings.get("prepare_data_only", False)))
        self.enable_llm_evaluation_var.set(
            bool(settings.get("enable_llm_evaluation", False))
        )
        self.skip_llm_editing_var.set(bool(settings.get("skip_llm_editing", False)))
        self.skip_quick_lstm_var.set(bool(settings.get("skip_quick_lstm", False)))
        self.skip_full_lstm_var.set(bool(settings.get("skip_full_lstm", False)))
        self.skip_batch_var.set(bool(settings.get("skip_batch", False)))
        self.no_lstm_cache_var.set(bool(settings.get("no_lstm_cache", False)))
        self.force_dataset_refresh_var.set(
            bool(settings.get("force_dataset_refresh", False))
        )
        self.quick_epochs_var.set(int(settings.get("quick_epochs", 1)))
        self.full_epochs_var.set(int(settings.get("full_epochs", 5)))
        self.experiment_runs_var.set(int(settings.get("experiment_runs", 1)))
        self.experiment_llm_runs_var.set(int(settings.get("experiment_llm_runs", 1)))
        self.poem_lines_var.set(int(settings.get("poem_lines", DEFAULT_POEM_LINES)))
        self.llm_provider_var.set(str(settings.get("llm_provider", "disabled")))

        self._refresh_preset_description()
        self._refresh_run_summary()
        if announce:
            self.status_var.set(f"Выбран сценарий: {preset['label']}")

    def _run_selected_preset(self) -> None:
        self._apply_selected_preset(announce=False)
        self._start_process()

    def _refresh_preset_description(self) -> None:
        preset = self._selected_preset()
        self.preset_description_var.set(f"Что делает сценарий: {preset['description']}")

    def _refresh_run_summary(self) -> None:
        poem_lines = self._safe_int(self.poem_lines_var, DEFAULT_POEM_LINES, minimum=2)
        quick_epochs = self._safe_int(self.quick_epochs_var, 1, minimum=1)
        full_epochs = self._safe_int(self.full_epochs_var, 5, minimum=1)
        experiment_runs = self._safe_int(self.experiment_runs_var, 1, minimum=1)
        experiment_llm_runs = self._safe_int(
            self.experiment_llm_runs_var,
            1,
            minimum=1,
        )

        if self.prepare_data_only_var.get():
            summary = (
                "Сейчас будет только подготовка данных: скачать датасет, собрать "
                "poems_clean.txt и разложить его по папкам проекта."
            )
            self.run_summary_var.set(summary)
            return

        steps = ["подготовка данных", "генерация Markov"]
        if not self.skip_quick_lstm_var.get():
            steps.append(f"quick LSTM ({quick_epochs} эпох.)")
        if not self.skip_full_lstm_var.get():
            steps.append(f"full LSTM ({full_epochs} эпох.)")
        if not self.skip_batch_var.get():
            steps.append(f"batch-сравнение ({experiment_runs} запуск.)")

        provider = self.llm_provider_var.get().strip().lower()
        if provider == "ollama" and not self.skip_llm_editing_var.get():
            llm_text = (
                f"LLM-редактирование через Ollama ({self.llm_model_name_var.get().strip() or DEFAULT_OLLAMA_MODEL})"
            )
        else:
            llm_text = "без LLM-редактирования"

        extra_notes = [
            f"стих на {poem_lines} строк по схеме {self.rhyme_scheme_var.get().strip() or DEFAULT_RHYME_SCHEME}",
            llm_text,
        ]
        if self.enable_llm_evaluation_var.get():
            extra_notes.append(
                f"LLM-оценка включена для batch ({experiment_llm_runs} запуск.)"
            )
        if self.no_lstm_cache_var.get():
            extra_notes.append("кэш LSTM будет проигнорирован")
        if self.force_dataset_refresh_var.get():
            extra_notes.append("датасет будет скачан заново")

        summary = (
            "Сейчас будет: "
            + ", ".join(steps)
            + ". Дополнительно: "
            + "; ".join(extra_notes)
            + "."
        )
        self.run_summary_var.set(summary)

    def _pick_dataset(self) -> None:
        selected = filedialog.askopenfilename(
            title="Выберите poems_clean.txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            initialdir=str(self.project_dir),
        )
        if selected:
            self.dataset_txt_var.set(selected)

    def _append_log(self, _kind: str, text: str) -> None:
        self.log_widget.configure(state="normal")
        self.log_widget.insert(tk.END, text)
        self.log_widget.see(tk.END)
        self.log_widget.configure(state="disabled")

    def _clear_log(self) -> None:
        self.log_widget.configure(state="normal")
        self.log_widget.delete("1.0", tk.END)
        self.log_widget.configure(state="disabled")
        self._append_log("info", "Лог очищен.\n")

    def _copy_command_preview(self) -> None:
        command = self._build_command()
        preview = subprocess.list2cmdline(command)
        self.root.clipboard_clear()
        self.root.clipboard_append(preview)
        self.status_var.set("Команда запуска скопирована в буфер обмена.")

    def _delete_lstm_weights(self) -> None:
        if self.process is not None:
            messagebox.showinfo(
                APP_TITLE,
                "Сейчас идёт запуск. Сначала остановите процесс, потом удаляйте веса.",
            )
            return

        message = (
            "Удалить сохранённые веса LSTM и локальный кэш?\n\n"
            "Будут удалены:\n"
            "- full LSTM: poet_model.keras / poet_metadata.json / poet_rhymes.json\n"
            "- quick LSTM: poet_quick_model.keras / poet_quick_metadata.json / poet_quick_rhymes.json\n"
            "- кэш из cache/poetry_lstm_cache\n\n"
            "После этого модель будет обучаться заново."
        )
        if not messagebox.askyesno(APP_TITLE, message):
            return

        lstm_dir = self.project_dir / "LSTM_generation_2"
        targets = [
            lstm_dir / "poet_model.keras",
            lstm_dir / "poet_metadata.json",
            lstm_dir / "poet_rhymes.json",
            lstm_dir / "poet_quick_model.keras",
            lstm_dir / "poet_quick_metadata.json",
            lstm_dir / "poet_quick_rhymes.json",
            self.cache_dir / "poet_model.keras",
            self.cache_dir / "poet_metadata.json",
            self.cache_dir / "poet_rhymes.json",
        ]

        deleted: List[str] = []
        missing: List[str] = []

        for path in targets:
            if path.exists():
                path.unlink()
                try:
                    deleted.append(str(path.relative_to(self.project_dir)))
                except Exception:
                    deleted.append(str(path))
            else:
                try:
                    missing.append(str(path.relative_to(self.project_dir)))
                except Exception:
                    missing.append(str(path))

        cache_dir_removed = False
        try:
            if self.cache_dir.exists() and not any(self.cache_dir.iterdir()):
                self.cache_dir.rmdir()
                cache_dir_removed = True
        except Exception:
            cache_dir_removed = False

        summary_lines = []
        if deleted:
            summary_lines.append("Удалено:")
            summary_lines.extend(f"- {item}" for item in deleted)
        else:
            summary_lines.append("Сохранённые веса не найдены.")

        if cache_dir_removed:
            summary_lines.append("- cache/poetry_lstm_cache")

        self.status_var.set(
            "LSTM-веса удалены." if deleted else "LSTM-веса не найдены."
        )
        self._append_log("info", "\n[LSTM RESET]\n" + "\n".join(summary_lines) + "\n")
        messagebox.showinfo(APP_TITLE, "\n".join(summary_lines))

    def _run_fast_mode(self) -> None:
        self.preset_var.set(PRESET_DEFINITIONS[0]["label"])
        self._run_selected_preset()

    def _run_full_mode(self) -> None:
        self.preset_var.set(PRESET_DEFINITIONS[1]["label"])
        self._run_selected_preset()

    def _run_prepare_only(self) -> None:
        self.preset_var.set(PRESET_DEFINITIONS[-1]["label"])
        self._run_selected_preset()

    def _build_command(self) -> List[str]:
        command = [
            str(self.python_executable),
            "-u",
            str(self.pipeline_path),
            "--project-dir",
            str(self.project_dir),
            "--start-line",
            self.start_line_var.get().strip() or DEFAULT_START_LINE,
            "--rhyme-scheme",
            self.rhyme_scheme_var.get().strip() or DEFAULT_RHYME_SCHEME,
            "--poem-lines",
            str(self._safe_int(self.poem_lines_var, DEFAULT_POEM_LINES, minimum=2)),
            "--quick-epochs",
            str(self._safe_int(self.quick_epochs_var, 1, minimum=1)),
            "--full-epochs",
            str(self._safe_int(self.full_epochs_var, 20, minimum=1)),
            "--experiment-runs",
            str(self._safe_int(self.experiment_runs_var, 1, minimum=1)),
            "--experiment-llm-runs",
            str(self._safe_int(self.experiment_llm_runs_var, 1, minimum=1)),
            "--llm-provider",
            self.llm_provider_var.get().strip() or "disabled",
        ]

        dataset_txt = self.dataset_txt_var.get().strip()
        if dataset_txt:
            command.extend(["--dataset-txt", dataset_txt])

        model_name = self.llm_model_name_var.get().strip()
        if model_name:
            command.extend(["--llm-model-name", model_name])

        flags = [
            (self.fast_mode_var.get(), "--fast"),
            (self.prepare_data_only_var.get(), "--prepare-data-only"),
            (self.enable_llm_evaluation_var.get(), "--enable-llm-evaluation"),
            (self.skip_llm_editing_var.get(), "--skip-llm-editing"),
            (self.skip_quick_lstm_var.get(), "--skip-quick-lstm"),
            (self.skip_full_lstm_var.get(), "--skip-full-lstm"),
            (self.skip_batch_var.get(), "--skip-batch"),
            (self.no_lstm_cache_var.get(), "--no-lstm-cache"),
            (self.force_dataset_refresh_var.get(), "--force-dataset-refresh"),
        ]
        for enabled, flag in flags:
            if enabled:
                command.append(flag)

        return command

    def _start_process(self) -> None:
        if self.process is not None:
            messagebox.showinfo(
                APP_TITLE,
                "Сейчас уже идёт запуск. Дождитесь завершения или нажмите 'Остановить'.",
            )
            return

        provider = self.llm_provider_var.get().strip().lower()
        ollama_base_url = self.ollama_base_url_var.get().strip()
        if provider == "ollama" and not ollama_base_url and not self.skip_llm_editing_var.get():
            messagebox.showwarning(
                APP_TITLE,
                "Для Ollama нужен URL локального сервера. Укажите адрес или переключите провайдер на disabled.",
            )
            return

        command = self._build_command()
        self._append_log("command", f"\n>>> {subprocess.list2cmdline(command)}\n")

        env = os.environ.copy()
        env.setdefault("PYTHONIOENCODING", "utf-8")
        env.pop("GROQ_API_KEY", None)
        if provider == "ollama" and ollama_base_url:
            env["OLLAMA_BASE_URL"] = ollama_base_url
        elif provider == "ollama":
            env.pop("OLLAMA_BASE_URL", None)

        creationflags = 0
        if sys.platform.startswith("win"):
            creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)

        self.process = subprocess.Popen(
            command,
            cwd=str(self.project_dir),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,
            env=env,
            creationflags=creationflags,
        )
        self.status_var.set("Пайплайн выполняется...")

        self.reader_thread = threading.Thread(
            target=self._read_process_output,
            daemon=True,
        )
        self.reader_thread.start()

    def _read_process_output(self) -> None:
        assert self.process is not None
        process = self.process
        if process.stdout is not None:
            for line in process.stdout:
                self.log_queue.put(("stdout", line))
        return_code = process.wait()
        self.log_queue.put(("done", str(return_code)))

    def _drain_log_queue(self) -> None:
        try:
            while True:
                kind, payload = self.log_queue.get_nowait()
                if kind in {"stdout", "info"}:
                    self._append_log(kind, payload)
                elif kind == "done":
                    self._handle_process_finished(int(payload))
        except queue.Empty:
            pass
        finally:
            self.root.after(100, self._drain_log_queue)

    def _handle_process_finished(self, return_code: int) -> None:
        if return_code == 0:
            self.status_var.set(
                "Пайплайн завершён успешно. Отчёты перечитаны и доступны во вкладках."
            )
            self._append_log("info", "\n[OK] Процесс завершён успешно.\n")
            self._refresh_reports()
        else:
            self.status_var.set(
                f"Пайплайн завершился с кодом {return_code}. Проверьте лог."
            )
            self._append_log(
                "info",
                f"\n[ERROR] Процесс завершился с кодом {return_code}.\n",
            )

        self.process = None
        self.reader_thread = None

    def _stop_process(self) -> None:
        if self.process is None:
            self.status_var.set("Активного процесса нет.")
            return
        self._append_log("info", "\n[STOP] Останавливаем процесс...\n")
        self.status_var.set("Останавливаем процесс...")
        self.process.terminate()

    def _refresh_hardware_async(self) -> None:
        self.nvidia_status_var.set("NVIDIA GPU: проверяем...")
        self.tensorflow_gpu_status_var.set("TensorFlow GPU: проверяем...")
        thread = threading.Thread(target=self._refresh_hardware_worker, daemon=True)
        thread.start()

    def _refresh_hardware_worker(self) -> None:
        nvidia_text = detect_nvidia_gpu_text()
        tensorflow_text = detect_tensorflow_gpu_text(self.python_executable)
        self.root.after(0, lambda: self.nvidia_status_var.set(nvidia_text))
        self.root.after(0, lambda: self.tensorflow_gpu_status_var.set(tensorflow_text))

    def _check_ollama_async(self) -> None:
        provider = self.llm_provider_var.get().strip().lower()
        if provider != "ollama":
            self.llm_status_var.set("Проверка Ollama доступна только когда провайдер = ollama.")
            return
        base_url = self.ollama_base_url_var.get().strip()
        if not base_url:
            self.llm_status_var.set("Ollama URL не задан.")
            return

        self.llm_status_var.set("Проверяем Ollama-соединение...")
        thread = threading.Thread(target=self._check_ollama_worker, daemon=True)
        thread.start()

    def _check_ollama_worker(self) -> None:
        model_name = self.llm_model_name_var.get().strip() or None
        base_url = self.ollama_base_url_var.get().strip()
        code = (
            "import json, os; "
            "from llm_poetry_tools import LLMPoetryAssistant; "
            "assistant = LLMPoetryAssistant("
            "provider='ollama', "
            "base_url=os.environ.get('OLLAMA_BASE_URL'), "
            f"model_name={model_name!r}"
            "); "
            "print(json.dumps(assistant.test_connection(), ensure_ascii=False))"
        )

        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["OLLAMA_BASE_URL"] = base_url

        try:
            completed = subprocess.run(
                [str(self.python_executable), "-c", code],
                check=True,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                env=env,
                cwd=str(self.project_dir),
            )
            payload = completed.stdout.strip()
            result = json.loads(payload) if payload else {}
            message = result.get("message", "Ollama status unavailable.")
            if result.get("ok"):
                response_text = str(result.get("response_text", "")).strip()
                if response_text:
                    message = f"{message} | probe={response_text}"
        except Exception as exc:
            message = f"Ollama probe failed: {exc}"

        self.root.after(0, lambda: self.llm_status_var.set(message))

    def _refresh_reports(self) -> None:
        self.detailed_reports = latest_report_files(self.reports_dir, "detailed")
        self.summary_reports = latest_report_files(self.reports_dir, "summary")

        detailed_values = [path.name for path in self.detailed_reports]
        summary_values = [path.name for path in self.summary_reports]

        self.detailed_report_combo["values"] = detailed_values
        self.detailed_reports_combo_reports_tab["values"] = detailed_values
        self.summary_report_combo["values"] = summary_values

        if detailed_values and self.detailed_report_var.get() not in detailed_values:
            self.detailed_report_var.set(detailed_values[0])
        if summary_values and self.summary_report_var.get() not in summary_values:
            self.summary_report_var.set(summary_values[0])

        self._load_selected_reports()

    def _selected_detailed_path(self) -> Optional[Path]:
        name = self.detailed_report_var.get().strip()
        if not name:
            return None
        path = self.reports_dir / name
        return path if path.exists() else None

    def _selected_summary_path(self) -> Optional[Path]:
        name = self.summary_report_var.get().strip()
        if not name:
            return None
        path = self.reports_dir / name
        return path if path.exists() else None

    def _load_selected_reports(self) -> None:
        detailed_path = self._selected_detailed_path()
        summary_path = self._selected_summary_path()

        self.detailed_rows = read_report_rows(detailed_path) if detailed_path else []
        self.summary_rows = read_report_rows(summary_path) if summary_path else []

        self.report_info_var.set(
            f"Подробный отчёт: {detailed_path if detailed_path else 'не выбран'} | "
            f"Краткий отчёт: {summary_path if summary_path else 'не выбран'} | "
            f"строк: detailed={len(self.detailed_rows)}, summary={len(self.summary_rows)}"
        )

        self._populate_poems_tree()
        self._populate_summary_tree()
        self._build_compare_pairs()

    def _populate_poems_tree(self) -> None:
        self.poems_tree.delete(*self.poems_tree.get_children())
        for index, row in enumerate(self.detailed_rows):
            self.poems_tree.insert(
                "",
                "end",
                iid=f"poem-{index}",
                values=(
                    row.get("report_scope", ""),
                    row.get("run", ""),
                    row.get("source_model", ""),
                    row.get("version", ""),
                    row.get("stage", ""),
                    row.get("actual_lines", row.get("lines", "")),
                    row.get("rhyme_quality", ""),
                    row.get("overall", ""),
                ),
            )

        self.poem_details_label.config(text="Выберите стих из списка.")
        self._set_text_widget(self.poem_text_widget, "")

    def _populate_summary_tree(self) -> None:
        self.summary_tree.delete(*self.summary_tree.get_children())
        for index, row in enumerate(self.summary_rows):
            self.summary_tree.insert(
                "",
                "end",
                iid=f"summary-{index}",
                values=(
                    row.get("report_scope", ""),
                    row.get("version", ""),
                    row.get("stage", ""),
                    row.get("row_count", ""),
                    row.get("source_model", ""),
                    row.get("mean_rhyme_quality", ""),
                    row.get("mean_overall", ""),
                ),
            )

    def _build_compare_pairs(self) -> None:
        grouped: Dict[tuple, Dict[str, Dict[str, str]]] = {}
        for row in self.detailed_rows:
            source_model = row.get("source_model", "").strip()
            if not source_model:
                continue
            key = (
                row.get("report_scope", ""),
                row.get("run", ""),
                source_model,
            )
            grouped.setdefault(key, {})
            stage = row.get("stage", "").strip() or "raw"
            if stage == "llm_edited":
                grouped[key]["llm"] = row
            else:
                grouped[key]["raw"] = row

        self.compare_pairs = []
        self.compare_tree.delete(*self.compare_tree.get_children())
        for index, (key, versions) in enumerate(grouped.items()):
            raw_row = versions.get("raw")
            llm_row = versions.get("llm")
            if not raw_row or not llm_row:
                continue
            pair = {
                "key": key,
                "raw": raw_row,
                "llm": llm_row,
            }
            self.compare_pairs.append(pair)
            self.compare_tree.insert(
                "",
                "end",
                iid=f"cmp-{index}",
                values=(key[0], key[1], key[2]),
            )

        self.compare_raw_label.config(text="Выберите пару raw/LLM слева.")
        self.compare_llm_label.config(text="Выберите пару raw/LLM слева.")
        self._set_text_widget(self.compare_raw_text, "")
        self._set_text_widget(self.compare_llm_text, "")

    def _on_poem_selected(self, _event: object) -> None:
        selection = self.poems_tree.selection()
        if not selection:
            return
        iid = selection[0]
        try:
            index = int(iid.split("-")[1])
        except Exception:
            return
        if not (0 <= index < len(self.detailed_rows)):
            return
        row = self.detailed_rows[index]
        self.poem_details_label.config(text=format_metrics_text(row))
        self._set_text_widget(self.poem_text_widget, row.get("poem_text", ""))

    def _on_compare_selected(self, _event: object) -> None:
        selection = self.compare_tree.selection()
        if not selection:
            return
        iid = selection[0]
        try:
            index = int(iid.split("-")[1])
        except Exception:
            return
        if not (0 <= index < len(self.compare_pairs)):
            return
        pair = self.compare_pairs[index]
        raw_row = pair["raw"]
        llm_row = pair["llm"]

        self.compare_raw_label.config(text=format_metrics_text(raw_row))
        self.compare_llm_label.config(text=format_metrics_text(llm_row))
        self._set_text_widget(self.compare_raw_text, raw_row.get("poem_text", ""))
        self._set_text_widget(self.compare_llm_text, llm_row.get("poem_text", ""))

    def _set_text_widget(self, widget: ScrolledText, text: str) -> None:
        widget.configure(state="normal")
        widget.delete("1.0", tk.END)
        widget.insert(tk.END, text or "")
        widget.configure(state="disabled")

    def _open_selected_detailed(self) -> None:
        path = self._selected_detailed_path()
        if not path:
            messagebox.showinfo(APP_TITLE, "Подробный отчёт пока не выбран.")
            return
        open_path(path)

    def _open_selected_summary(self) -> None:
        path = self._selected_summary_path()
        if not path:
            messagebox.showinfo(APP_TITLE, "Краткий отчёт пока не выбран.")
            return
        open_path(path)

    def _save_selected_detailed_copy(self) -> None:
        self._save_report_copy(self._selected_detailed_path())

    def _save_selected_summary_copy(self) -> None:
        self._save_report_copy(self._selected_summary_path())

    def _save_report_copy(self, source_path: Optional[Path]) -> None:
        if not source_path:
            messagebox.showinfo(APP_TITLE, "Сначала выберите report.")
            return
        suffix = source_path.suffix.lower() or ".xlsx"
        if suffix == ".xlsx":
            filetypes = [("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")]
        else:
            filetypes = [("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")]
        target = filedialog.asksaveasfilename(
            title="Сохранить копию отчёта",
            initialfile=source_path.name,
            defaultextension=suffix,
            filetypes=filetypes,
        )
        if not target:
            return
        shutil.copy2(source_path, target)
        self.status_var.set(f"Копия отчёта сохранена: {target}")

    def _on_close(self) -> None:
        if self.process is not None:
            if not messagebox.askyesno(
                APP_TITLE,
                "Сейчас идёт запуск. Закрыть окно и остановить процесс?",
            ):
                return
            self.process.terminate()
        self.root.destroy()


def main() -> int:
    root = tk.Tk()
    style = ttk.Style(root)
    if "vista" in style.theme_names():
        style.theme_use("vista")
    PipelineDesktopApp(root)
    try:
        root.mainloop()
    except KeyboardInterrupt:
        try:
            root.destroy()
        except Exception:
            pass
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
